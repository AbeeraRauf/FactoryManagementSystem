import datetime
import gc
import re
import sys
import traceback
from datetime import date, datetime
from math import ceil

import matplotlib.pyplot as plt
 
import openpyxl
import pandas as pd
import PyQt5
import pywhatkit
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pandas import DataFrame, Series
from pdf2image import convert_from_path
from PIL import Image, ImageTk
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QDate, QEvent
from PyQt5.QtGui import QDoubleValidator, QValidator
from PyQt5.QtWidgets import (QMessageBox, QTableWidget, QTableWidgetItem,
                             QWidget)
from PyQt5.uic import loadUi

from cashrecieptgenerator import generate_cash_reciept
from customerrecieptgenerator import generate_customer_invoice

details=[]
reelsstock = pd.read_excel(r'book.xlsx', index_col=None, usecols=['date','Item_Type', 'Size', 'Weight_g','vendor','rate'],sheet_name='reels_stock')
reelsstock['Weight_g']=reelsstock['Weight_g'].astype(int)
vendorg=[]
 
CashCustomers = pd.read_excel('book.xlsx', index_col=None, sheet_name='cash table',  usecols=['DATE', 'RECIEPT_NUMBER', 'CLIENT_ID', 'CLIENT_NAME',
                                                       'CONTACT_NO', 'DETAILS_OF_BILL', 'DEBIT', 'CREDIT',
                                                       'CREDIT_DETAILS', 'RENT', 'BALANCE'])
pdf_date = ''
pdf_client_name = ''
pdf_client_contact_number = ''
pdf_reciept_number = ''
pdf_debit = 0.0
pdf_credit = 0.0
pdf_credit_details = ''
pdf_rent = 0.0
pdf_previous_balance = 0.0
pdf_previous_bill_number = ''
pdf_total_balance = 0.0
pdf_dataframe = pd.DataFrame()
 
class Ui_Form(object):
    
    def setupUi6(self, Form):
        Form.setObjectName("Form")
        Form.setFixedSize(1360, 850)
        Form.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.tableWidget = QTableWidget(Form)
        
        self.tableWidget.setGeometry(QtCore.QRect(25, 410, 1255, 371))
        
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(4)
        self.tableWidget.setRowCount(0)   
        self.tableWidget.setHorizontalHeaderLabels(('CLIENT_ID', 'Name','Contact','BALANCE'))  # set header text
        header = self.tableWidget.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
        self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.groupBox = QtWidgets.QGroupBox(Form)
        self.groupBox.setObjectName(u"groupBox")
        self.groupBox.setGeometry(QtCore.QRect(25, 120, 531, 271))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.groupBox.setFont(font)
        self.groupBox.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.searchbyname = QtWidgets.QTextEdit(self.groupBox)
        self.searchbyname.setObjectName(u"searchbyname")
        self.searchbyname.setEnabled(False)
        self.searchbyname.setGeometry(QtCore.QRect(300, 120, 191, 31))
        self.searchbyname.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.searchbyid = QtWidgets.QTextEdit(self.groupBox)
        self.searchbyid.setObjectName(u"searchbyid")
        self.searchbyid.setGeometry(QtCore.QRect(300, 60, 191, 31))
        self.searchbyid.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.searchbutton = QtWidgets.QPushButton(self.groupBox)
        self.searchbutton.setObjectName(u"searchbutton")
        self.searchbutton.setGeometry(QtCore.QRect(180, 200, 151, 51))
        font1 = QtGui.QFont()
        font1.setPointSize(12)
        font1.setBold(True)
        font1.setWeight(75)
        self.searchbutton.setFont(font1)
        self.searchbutton.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        font11 = QtGui.QFont()
        font11.setPointSize(14)
        font11.setBold(True)
        font11.setWeight(75)
        self.label_2 = QtWidgets.QRadioButton(self.groupBox)
        self.label_2.setObjectName(u"label_2")
        self.label_2.setGeometry(QtCore.QRect(50, 70, 201, 31))
        self.label_2.setFont(font11)
        self.label_2.setChecked(True)
        self.label_3 = QtWidgets.QRadioButton(self.groupBox)
        self.label_3.setObjectName(u"label_3")
        
        self.label_3.setGeometry(QtCore.QRect(44, 120, 231, 31))
        self.label_3.setFont(font11)
        
        self.addnewbill = QtWidgets.QPushButton(Form)
        self.addnewbill.setObjectName(u"addnewbill")
        self.addnewbill.setGeometry(QtCore.QRect(800, 790, 140,41))
        self.addnewbill.setFont(font1)
        self.addnewbill.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.addclientsgroupbox = QtWidgets.QGroupBox(Form)
        self.addclientsgroupbox.setObjectName(u"addclientsgroupbox")
        self.addclientsgroupbox.setEnabled(True)
        self.addclientsgroupbox.setGeometry(QtCore.QRect(675, 120, 605, 271))
        self.addclientsgroupbox.setFont(font)
        self.addclientsgroupbox.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        
        
         
        self.name = QtWidgets.QLabel(self.addclientsgroupbox)
        self.name.setObjectName(u"name")
        self.name.setGeometry(QtCore.QRect(10, 30, 70, 61))
        self.name.setFont(font11)
        self.name_2 = QtWidgets.QTextEdit(self.addclientsgroupbox)
        self.name_2.setObjectName(u"name_2")
        self.name_2.setGeometry(QtCore.QRect(115, 40, 161, 41))
        self.name_2.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.contact_2 = QtWidgets.QTextEdit(self.addclientsgroupbox)
        self.contact_2.setObjectName(u"contact_2")
        self.contact_2.setGeometry(QtCore.QRect(430, 40, 161, 41))
        self.contact_2.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.contact = QtWidgets.QLabel(self.addclientsgroupbox)
        self.contact.setObjectName(u"contact")
        self.contact.setGeometry(QtCore.QRect(300, 30, 90, 61))
        self.contact.setFont(font11)
        self.l1 = QtWidgets.QLabel(self.addclientsgroupbox)
        self.l1.setObjectName(u"l1")
        self.l1.setGeometry(QtCore.QRect(300, 100, 100, 31))
        self.l1.setFont(font11)
        self.l2_lb = QtWidgets.QLabel(self.addclientsgroupbox)
        self.l2_lb.setObjectName(u"l2_lb")
        self.l2_lb.setGeometry(QtCore.QRect(300, 130, 100, 31))
        self.l2_lb.setFont(font11)
        self.l3 = QtWidgets.QLabel(self.addclientsgroupbox)
        self.l3.setObjectName(u"l3")
        self.l3.setGeometry(QtCore.QRect(300, 158, 100, 31))
        self.l3.setFont(font11)
        self.m_ledger = QtWidgets.QTextEdit(self.addclientsgroupbox)
        self.m_ledger.setObjectName(u"m_ledger")
        self.m_ledger.setGeometry(QtCore.QRect(430,110,161,41))
        self.m_ledger.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.m_ledger.setText("0")
        self.manualledger = QtWidgets.QLabel(self.addclientsgroupbox)
        self.manualledger.setObjectName(u"balance")
        self.manualledger.setGeometry(QtCore.QRect(10, 100, 92, 61))
        self.manualledger.setFont(font11)
        self.balance_tf = QtWidgets.QTextEdit(self.addclientsgroupbox)
        self.balance_tf.setObjectName(u"m_ledger_2")
        self.balance_tf.setGeometry(QtCore.QRect(115,110,161,41))
        self.balance_tf.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.balance_tf.setText("0")
        self.addnewclient = QtWidgets.QPushButton(self.addclientsgroupbox)
        self.addnewclient.setObjectName(u"addnewclient")
        self.addnewclient.setGeometry(QtCore.QRect(220, 200, 161, 51))
        self.addnewclient.setFont(font1)
        self.addnewclient.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.resettable = QtWidgets.QPushButton(Form)
        self.resettable.setObjectName(u"resettable")
        self.resettable.setGeometry(QtCore.QRect(600, 790, 131, 41))
        self.resettable.setFont(font1)
        self.resettable.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.textBrowser = QtWidgets.QTextBrowser(Form)
        self.textBrowser.setObjectName(u"textBrowser")
        self.textBrowser.setEnabled(False)
        self.textBrowser.setGeometry(QtCore.QRect(0, 10, 1360, 85))
        self.textBrowser.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"gridline-color: rgb(0, 0, 127);\n"
"color:rgb(255,255, 255) ;")
        self.textBrowser_3 = QtWidgets.QTextBrowser(Form)
        self.textBrowser_3.setObjectName(u"textBrowser_3")
        self.textBrowser_3.setEnabled(False)
        self.textBrowser_3.setGeometry(QtCore.QRect(10, 960, 1901, 31))
        self.textBrowser_3.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"gridline-color: rgb(0, 0, 127);\n"
"color:rgb(255,255, 255) ;") 
        
         
        
        self.del_entry_btn = QtWidgets.QPushButton(Form)
        self.del_entry_btn.setGeometry(QtCore.QRect(400, 790, 140, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.del_entry_btn.setFont(font)
        self.del_entry_btn.setStyleSheet("background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.del_entry_btn.setObjectName("del_entry_btn")
        


        self.printbutton = QtWidgets.QPushButton(Form)
        self.printbutton.setGeometry(QtCore.QRect(1148, 790, 131, 41))
        self.printbutton.setObjectName("printbutton")
        self.printbutton.setText("Print")
        self.printbutton.setFont(font1)
        self.printbutton.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.printbutton.setVisible(True)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)
        
        
        
    def updatetable(self):
            self.tableWidget.clear ()
            self.tableWidget.setHorizontalHeaderLabels(('CLIENT_ID', 'Name','Contact','BALANCE','Page on Manual Ledger'))  # set header text
            self.tableWidget.setObjectName("tableWidget")
            self.tableWidget.setColumnCount(5)
            self.tableWidget.setRowCount(0)   
            self.tableWidget.setHorizontalHeaderLabels(('CLIENT_ID', 'Name','Contact','BALANCE','Page on Manual Ledger'))  # set header text
            header = self.tableWidget.horizontalHeader()       
            header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
            header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
            header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
            header.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
            header.setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
            self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
            clientsallsheets=pd.read_excel('book.xlsx', index_col=None,sheet_name=None)
            toremove1=['reels_stock_in_out','tota_stock_in_out','rolls_stock_in_out','client info','rolls_stock','reels_stock','totay','Fluting', 'Fluting_Bareek', 'L1', 'L1_Bareek', 'L2', 'L2_Bareek', 'Test_Liner', 'Test_Liner_Bareek', 'Box_Board_2_5_No', 'Box_Board_2_5_No_Bareek', 'Box_Board_3_No', 'Box_Board_3_No_Bareek', 'Local_Kraft', 'Local_Kraft_Bareek', 'Imported_Kraft', 'Imported_Kraft_Bareek', 'Super_Fluting', 'Super_Fluting_Bareek','cash table']
 
            sheets1=[]
            for i in  clientsallsheets.keys():
                if i not in toremove1:
                    sheets1.append(i) 
            datatodisplay2=pd.read_excel('book.xlsx', index_col=None,sheet_name=sheets1,usecols = ['CLIENT_ID','CLIENT_NAME','CONTACT_NO','BALANCE','Page_on_Manual_Ledger'])

            data1=[]
            try:
                    for i in datatodisplay2.keys():

                        data1.append(datatodisplay2[i].values[-1]) 

                    datatodisplay=pd.DataFrame(data=data1,columns = ['CLIENT_ID','CLIENT_NAME','CONTACT_NO','BALANCE','Page_on_Manual_Ledger']) 
                    datatodisplay=datatodisplay.sort_values('CLIENT_ID')

                    rowPosition = 0    
                    for row in datatodisplay.iterrows():

                        self.tableWidget.insertRow(rowPosition)
                        self.tableWidget.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(str(row[1][0])))
                        self.tableWidget.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(str(row[1][1])))
                        self.tableWidget.setItem(rowPosition, 2, QtWidgets.QTableWidgetItem(str(row[1][2])))
                        self.tableWidget.setItem(rowPosition, 3, QtWidgets.QTableWidgetItem(str(row[1][3])))
                        self.tableWidget.setItem(rowPosition, 4, QtWidgets.QTableWidgetItem(str(row[1][4])))
                        rowPosition+=1
            except IndexError:
                pass
            
            self.label_3.setChecked(False)
            self.label_2.setChecked(True)
            self.searchbyname.setEnabled(False)
            self.searchbyid.setEnabled(True)
            self.searchbyname.setText("")
            self.searchbyid.setText("")
            self.m_ledger.setText('0') 
            self.contact_2.setText('') 
            self.name_2.setText('')
       
    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Clients"))
         
        self.groupBox.setTitle(QtCore.QCoreApplication.translate("Form", u"Search", None))
        self.searchbutton.setText(QtCore.QCoreApplication.translate("Form", u"Search", None))
        self.label_2.setText(QtCore.QCoreApplication.translate("Form", u"Search By ID:", None))
        self.label_3.setText(QtCore.QCoreApplication.translate("Form", u"Search By Name:", None))
        self.addnewbill.setText(QtCore.QCoreApplication.translate("Form", u"Add New Bill", None))
        self.addclientsgroupbox.setTitle(QtCore.QCoreApplication.translate("Form", u"Add New Client", None))
        self.name.setText(QtCore.QCoreApplication.translate("Form", u"Name:", None))
        self.name_2.setPlaceholderText(QtCore.QCoreApplication.translate("Form", u"FirstName SecondName SurName", None))
        self.contact_2.setPlaceholderText(QtCore.QCoreApplication.translate("Form", u"xxxxxxxxxxx", None))
        self.contact.setText(QtCore.QCoreApplication.translate("Form", u"Contact:", None))
        self.l1.setText(QtCore.QCoreApplication.translate("Form", u"Page On", None))
        self.l2_lb.setText(QtCore.QCoreApplication.translate("Form", u"Manual", None))
        self.l3.setText(QtCore.QCoreApplication.translate("Form", u"Ledger", None))
        self.m_ledger.setPlaceholderText(QtCore.QCoreApplication.translate("Form", u"0", None))
        self.manualledger.setText(QtCore.QCoreApplication.translate("Form", u"Balance:", None))
        self.addnewclient.setText(QtCore.QCoreApplication.translate("Form", u"Add New Client", None))
        self.resettable.setText(QtCore.QCoreApplication.translate("Form", u"Refresh", None))
        self.textBrowser.setHtml(QtCore.QCoreApplication.translate("Form", u"<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:'MS Shell Dlg 2'; font-size:7.8pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:36pt; font-weight:600;\">  Ahmed Corrugation Machines</span></p></body></html>", None))
        self.textBrowser_3.setHtml(QtCore.QCoreApplication.translate("Form", u"<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:'MS Shell Dlg 2'; font-size:7.8pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>", None))
        self.del_entry_btn.setText(_translate("Form", "Delete Entry"))

        self.updatetable()
        
        def close_window():
            QtCore.QTimer.singleShot(0, Form.close)
        
        def write_excel(df, sheets, excel_path):
            book=openpyxl.load_workbook(excel_path)
            writer = pd.ExcelWriter(excel_path, engine='openpyxl',mode="a",if_sheet_exists="replace")
            writer.book = book
            writer.sheets = {ws.title:ws for ws in book.worksheets}
            result=pd.DataFrame()
            result = df
            result.to_excel(writer,sheet_name=sheets, index=False)
            writer.save() 
        
        def addnewclient(): 
            name=self.name_2.toPlainText().strip()
            contact=self.contact_2.toPlainText().strip()
             
            ledger=self.m_ledger.toPlainText().strip()
            
            balance=self.balance_tf.toPlainText()
            if balance=='':
                balance='0'
            
            errors=[]       
            if (contact == '') or (not bool(re.match('^[1-9]\d{11,12}$' , contact))):
                errors.append('Wrong Input in Client Contact Text Field')
                  
            if     (not bool(re.match("\d+$",ledger))):
                errors.append('Wrong Input in Client balance Text Field ') 
                  
                    
            if (not (bool(re.match("^[a-zA-z]+([\s][a-zA-Z]+)*$",name)))) or (name=='') :
                errors.append('Wrong Input in Client name Text Field ')
            
            
            if (len(errors) != 0  ):
                 
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs=" , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text
                
                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message=msg.exec_()
            
            if(len(errors)==0):
                clientssheets=pd.read_excel('book.xlsx', index_col=None,sheet_name=None)
                toremove=['reels_stock_in_out','tota_stock_in_out','rolls_stock_in_out','client info','rolls_stock','reels_stock','totay','Fluting', 'Fluting_Bareek', 'L1', 'L1_Bareek', 'L2', 'L2_Bareek', 'Test_Liner', 'Test_Liner_Bareek', 'Box_Board_2_5_No', 'Box_Board_2_5_No_Bareek', 'Box_Board_3_No', 'Box_Board_3_No_Bareek', 'Local_Kraft', 'Local_Kraft_Bareek', 'Imported_Kraft', 'Imported_Kraft_Bareek', 'Super_Fluting', 'Super_Fluting_Bareek','cash table']
                sheets=[]
                for i in clientssheets.keys():
                    if i not in toremove:
                        sheets.append(i)
                clientsdata=pd.read_excel('book.xlsx', index_col=None,sheet_name=sheets)
                #clientsalldata=pd.Dataframe(data=clientsdata)
                datatodisplay1=pd.read_excel('book.xlsx', index_col=None,sheet_name=sheets,usecols = ['CLIENT_ID','CLIENT_NAME','CONTACT_NO','BALANCE','Page_on_Manual_Ledger'])
                data1=[]
                for i in datatodisplay1.keys():
                    data1.append(datatodisplay1[i].values[-1])
                datatodisplay=pd.DataFrame(data=data1,columns = ['CLIENT_ID','CLIENT_NAME','CONTACT_NO','BALANCE','Page_on_Manual_Ledger']) 
                datatodisplay=datatodisplay.sort_values('CLIENT_ID')
                idslist=[] 
                for i in sheets:
                    idslist.append(int(i))
                if(len(idslist)==0):
                    CLIENT_ID=1
                else:
                    idslist.sort()
                    client_id=idslist[-1]+1
                    CLIENT_ID=client_id
                sheetname=str(CLIENT_ID)
                DATE=str(date.today().strftime("%d-%m-%y") )
                
                
                 
                
                RECIEPT_NUMBER=0
                CLIENT_NAME=name
                CONTACT_NO=contact
                DETAILS_OF_BILL=''
                DEBIT=0
                CREDIT=0
                CREDIT_DETAILS=0
                RENT=0
                BALANCE=balance
                Ledger=ledger
             # dataframe Name and Age columns
                #global clientsallsheetsdata
                df= pd.DataFrame({'DATE':DATE ,'RECIEPT_NUMBER':RECIEPT_NUMBER ,'CLIENT_ID':CLIENT_ID,'CLIENT_NAME':CLIENT_NAME,'CONTACT_NO':CONTACT_NO,'DETAILS_OF_BILL':DETAILS_OF_BILL,'DEBIT':DEBIT,'CREDIT':CREDIT,'CREDIT_DETAILS':CREDIT_DETAILS,'RENT':RENT,'BALANCE':BALANCE,'Page_on_Manual_Ledger':ledger},index=[0])
                write_excel(df,sheetname,'book.xlsx')
                self.updatetable()
                self.label_3.setChecked(False)
                self.label_2.setChecked(True)
                self.searchbyname.setEnabled(False)
                self.searchbyid.setEnabled(True)
                self.searchbyname.setText("")
                self.searchbyid.setText("")
                self.m_ledger.setText('0') 
                self.contact_2.setText('') 
                self.name_2.setText('') 
        
        def searchuser():
            #label_3=name label_2=id
            
            errors=[]
            name=self.searchbyname.toPlainText().strip()
            ids=self.searchbyid.toPlainText().strip()
            
            if ((self.label_2.isChecked()==True) and (self.label_3.isChecked()==False)):
                if((not (bool(re.match("\d+$",ids))))) :
                    errors.append('Wrong Input in Client ID Text Field ') 
                else:
                    columnOfInterest =0 # or whatever
                    valueOfInterest =ids 
                    for rowIndex in range(self.tableWidget.rowCount()):
                        twItem = self.tableWidget.item(rowIndex, columnOfInterest)
                        if (valueOfInterest)== twItem.text().strip() :
                            self.tableWidget.setRowHidden(rowIndex, False)
                        else:
                            self.tableWidget.setRowHidden(rowIndex, True)

            elif(self.label_3.isChecked()==True  and (self.label_2.isChecked()==False)):
                if((not (bool(re.match("^[a-zA-z]+([\s][a-zA-Z]+)*$",name))))  ) :
                    errors.append('Wrong Input in Client Name Text Field ') 
                else:
                    columnOfInterest =1 # or whatever
                    valueOfInterest = name 
                    for rowIndex in range(self.tableWidget.rowCount()):
                        twItem = self.tableWidget.item(rowIndex, columnOfInterest)
                        if (valueOfInterest.lower()) in twItem.text().strip().lower() :
                            self.tableWidget.setRowHidden(rowIndex, False)
                        else:
                            self.tableWidget.setRowHidden(rowIndex, True)
            else:
                errors.append("Select Search by Id / Search by Name to do search")  # set text
                
                 
            if (len(errors) != 0  ):
                 
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs=" , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text
                msg.setWindowTitle("Alert")  # set title
                message=msg.exec_()        
                    
                
                
            
        def addnewbills():
                        current_row = self.tableWidget.currentRow()
                        try:  
                            CLIENTID1=self.tableWidget.item(current_row,0).text().strip()
                            CLIENT_ID=int(float(CLIENTID1))
                            CLIENT_NAME=self.tableWidget.item(current_row,1).text().strip()
                            CONTACT_NO=self.tableWidget.item(current_row,2).text().strip()
                            BALANCE=self.tableWidget.item(current_row,3).text().strip()
                            LEDGER=self.tableWidget.item(current_row,4).text()  .strip()  
                            datatodisplay1=pd.read_excel('book.xlsx' ,sheet_name=CLIENTID1)
                            datatodisplay1=datatodisplay1.iloc[-1:]
                        
                         
                            try:
                                x=datatodisplay1['RECIEPT_NUMBER'].values[0]

                                result=x
                            except:
                                 result=0


                            clientrow=[str(CLIENT_ID),str(CLIENT_NAME),CONTACT_NO,BALANCE,LEDGER,result]
                            #import ClientBill as clientbill
                           

                            try:        
                                clientbillOBJ=Ui_MainWindow_client(clientrow)
                                MainWindow = QtWidgets.QMainWindow()
                                clientbillOBJ.setupUi(MainWindow)
                                MainWindow.showMaximized()

                            except AttributeError:
                                         traceback.print_exc()
                        except:   
                                   pass           
                                    
                                    
        def byidsearch():
            #print("select by id name")
            self.searchbyname.setEnabled(False)
            self.searchbyid.setEnabled(True)
            self.searchbyname.setText("")
        def bynamesearch():
            #print("select by name")
            self.searchbyname.setEnabled(True)
            self.searchbyid.setEnabled(False)
            self.searchbyid.setText("")
            
        def del_entry():
                    dates=''
                    current_row = self.tableWidget.currentRow()
                    current_column = self.tableWidget.currentColumn()
                    cell_value = self.tableWidget.item(current_row, current_column).text().strip() 
                    #print(cell_value[0] )
            
                    try:
                        
                        workbook=openpyxl.load_workbook('book.xlsx')
                        del workbook[cell_value]
                        workbook.save('book.xlsx')
                        self.tableWidget.removeRow(self.tableWidget.currentRow())
                    except KeyError:
                        traceback.print_exc()   

                    finally:
                        msg = QMessageBox()  # create an instance of it
                        msg.setIcon(QMessageBox.Information)  # set icon
                        
                        msg.setText("Wrong Selection. Select whole row or first cell of row you want to delete.")  # set text

                        '''msg.setInformativeText()'''  # set information under the main text
                        msg.setWindowTitle("Alert")  # set title
                        message = msg.exec_()
            
             
        def generate_customer_pdf():
          
            rowscount=self.tableWidget.rowCount()
            headercount =  self.tableWidget.columnCount()
            # headertext =  self.tableWidget.horizontalHeaderItem(x).text()

            table=pd.DataFrame(columns=[self.tableWidget.horizontalHeaderItem(i).text() for i in range(headercount)],
                            index=[x for x in range(rowscount)])

            for row in range(rowscount):
                for col in range(headercount):
                    headertext =  self.tableWidget.horizontalHeaderItem(col).text()
                    cell =  self.tableWidget.item(row, col).text()  # get cell at row, col
                    table[headertext][row]=cell
                    

            #table.to_csv('C:\\Users\\Hp\\Desktop\\Factory_management_system\\table.csv')
            table.to_csv('table.csv')
            #print('\n\nList',mylist)
            generate_customer_invoice(table)

        self.del_entry_btn.clicked.connect(del_entry)  
        self.addnewbill.clicked.connect(addnewbills)
        self.addnewclient.clicked.connect(addnewclient)
        self.searchbutton.clicked.connect(searchuser)
        self.resettable.clicked.connect(self.updatetable)
        self.label_3.clicked.connect(bynamesearch)
        self.label_2.clicked.connect(byidsearch)
        self.printbutton.clicked.connect(generate_customer_pdf)
    
    
class Ui_MainWindow_client(object):
    def __init__(self, clientrecord):
        
        self.clientrecord=clientrecord
        self.clients=pd.read_excel('book.xlsx', index_col=None,sheet_name=self.clientrecord[0])
        self.reelsstock = pd.read_excel(r'book.xlsx', index_col=None, usecols=['date','Item_Type', 'Size', 'Weight_g','vendor','rate'],sheet_name='reels_stock')
        self.reelsstock['Weight_g']=self.reelsstock['Weight_g'].astype(int)
        
        self.totaystock = pd.read_excel(r'book.xlsx', index_col=None, usecols=['date','Item_Type', 'Size', 'Weight_g','detail','rate'],sheet_name='totay')
        self.totaystock['Weight_g']=self.totaystock['Weight_g'].astype(int)

        self.rollsstock = pd.read_excel(r'book.xlsx', index_col=None, usecols=['ID','Type','Rate','Size','Description','Quantity'],sheet_name='rolls_stock')
        self.Fluting = pd.read_excel(r'book.xlsx', index_col=None, usecols=['FLUTINGID','Size','Quantity'],sheet_name='Fluting')
        self.Fluting_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['FLUTINGBID','Size','Quantity'],sheet_name='Fluting_Bareek')
        self.L1 = pd.read_excel(r'book.xlsx', index_col=None, usecols=['L1ID','Size','Quantity'],sheet_name='L1')
        self.L1_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['L1BID','Size','Quantity'],sheet_name='L1_Bareek')
        self.L2 = pd.read_excel(r'book.xlsx', index_col=None, usecols=['L2ID','Size','Quantity'],sheet_name='L2')
        self.L2_Bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['L2BID','Size','Quantity'],sheet_name='L2_Bareek')
        self.testliner = pd.read_excel(r'book.xlsx', index_col=None, usecols=['TLID','Size','Quantity'],sheet_name='Test_Liner')
        self.testliner_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['TLBID','Size','Quantity'],sheet_name='Test_Liner_Bareek')
        self.boxboard2_5 = pd.read_excel(r'book.xlsx', index_col=None, usecols=['BB25ID','Size','Quantity'],sheet_name='Box_Board_2_5_No')
        self.boxboard2_5_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['BB25BID','Size','Quantity'],sheet_name='Box_Board_2_5_No_Bareek')
        self.boxboard3 = pd.read_excel(r'book.xlsx', index_col=None, usecols=['BB3ID','Size','Quantity'],sheet_name='Box_Board_3_No')
        self.boxboard3_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['BB3BID','Size','Quantity'],sheet_name='Box_Board_3_No_Bareek')
        self.localkraft = pd.read_excel(r'book.xlsx', index_col=None, usecols=['LKID','Size','Quantity'],sheet_name='Local_Kraft')
        self.localkraft_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['LKBID','Size','Quantity'],sheet_name='Local_Kraft_Bareek')
        self.importedkraft = pd.read_excel(r'book.xlsx', index_col=None, usecols=['KID','Size','Quantity'],sheet_name='Imported_Kraft')
        self.importedkraft_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['KBID','Size','Quantity'],sheet_name='Imported_Kraft_Bareek')
        self.Super_Fluting = pd.read_excel(r'book.xlsx', index_col=None, usecols=['SFID','Size','Quantity'],sheet_name='Super_Fluting')
        self.Super_Fluting_bareek = pd.read_excel(r'book.xlsx', index_col=None, usecols=['SFBID','Size','Quantity'],sheet_name='Super_Fluting_Bareek')
        self.rollsquantitylist = [self.Fluting['Quantity'].sum(skipna=True),]
        a = list()
        a.append(self.Fluting['Quantity'].sum(skipna=True))
        a.append(self.Fluting_bareek['Quantity'].sum(skipna=True))
        a.append(self.L1['Quantity'].sum(skipna=True))
        a.append(self.L1_bareek['Quantity'].sum(skipna=True))
        a.append(self.L2['Quantity'].sum(skipna=True))
        a.append(self.L2_Bareek['Quantity'].sum(skipna=True))
        a.append(self.testliner['Quantity'].sum(skipna=True))
        a.append(self.testliner_bareek['Quantity'].sum(skipna=True))
        a.append(self.boxboard2_5['Quantity'].sum(skipna=True))
        a.append(self.boxboard2_5_bareek['Quantity'].sum(skipna=True))
        a.append(self.boxboard3['Quantity'].sum(skipna=True))
        a.append(self.boxboard3_bareek['Quantity'].sum(skipna=True))
        a.append(self.localkraft['Quantity'].sum(skipna=True))
        a.append(self.localkraft_bareek['Quantity'].sum(skipna=True))
        a.append(self.importedkraft['Quantity'].sum(skipna=True))
        a.append(self.importedkraft_bareek['Quantity'].sum(skipna=True))
        a.append(self.Super_Fluting['Quantity'].sum(skipna=True))
        a.append(self.Super_Fluting_bareek['Quantity'].sum(skipna=True))
        self.rollsstock['Quantity']=a
        self.stock_out_rolls = pd.read_excel(r'book.xlsx', index_col=None, usecols=['date','details','item_type','size','quantity','quantity_in_stock'],sheet_name='rolls_stock_in_out')
        self.stock_out_reels = pd.read_excel(r'book.xlsx', index_col=None, usecols=['date','details','item_type','size','weight','rate'],sheet_name='reels_stock_in_out')
        self.stock_out_totay = pd.read_excel(r'book.xlsx', index_col=None, usecols=['date','details','item_type','size','weight','rate'],sheet_name='tota_stock_in_out')
        

    

    def setupUi(self, MainWindow ):
        #timenow = datetime.datetime.now()
        MainWindow.setObjectName("Client Bill")
        MainWindow.setFixedSize(1460,850)
        MainWindow.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.itemListComboBox = QtWidgets.QComboBox(self.centralwidget)
        self.itemListComboBox.setGeometry(QtCore.QRect(840, 197, 231, 32))
        self.itemListComboBox.setObjectName("itemListComboBox")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        self.itemListComboBox.addItem("")
        font = QtGui.QFont()
        font.setFamily(u"MS Shell Dlg 2")
        self.itemListComboBox.setFont(font)
        self.item_label = QtWidgets.QLabel(self.centralwidget)
        self.item_label.setObjectName(u"item_label")
        self.item_label.setGeometry(QtCore.QRect(750, 194, 61, 41))
        font1 = QtGui.QFont()
        font1.setFamily(u"MS Shell Dlg 2")
        font1.setPointSize(12)
        font1.setBold(True)
        font1.setWeight(75)
        self.item_label.setFont(font1)
        self.item_label.setTextFormat(QtCore.Qt.RichText)
        self.item_label.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label.setWordWrap(True)
        
        font2 = QtGui.QFont()
        font2.setPointSize(12)
        font2.setBold(True)
        font2.setWeight(75)  
        
        self.RollsContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.RollsContainer.setObjectName(u"RollsContainer")
        self.RollsContainer.setGeometry(QtCore.QRect(752, 240, 671, 301))
        self.RollsContainer.setFont(font2)
        self.RollsContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.RollsPaperTypeComboBox = QtWidgets.QComboBox(self.RollsContainer)
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        self.RollsPaperTypeComboBox.addItem("")
        
        self.RollsPaperTypeComboBox.setObjectName(u"RollsPaperTypeComboBox")
        self.RollsPaperTypeComboBox.setGeometry(QtCore.QRect(100, 40, 191, 31))
        self.RollsPaperTypeComboBox.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_3 = QtWidgets.QLabel(self.RollsContainer)
        self.item_label_3.setObjectName(u"item_label_3")
        self.item_label_3.setGeometry(QtCore.QRect(30, 30, 51, 51))
        font3 = QtGui.QFont()
        font3.setFamily(u"MS Shell Dlg 2")
        font3.setPointSize(11)
        font3.setBold(True)
        font3.setWeight(75)
        
        
        font12 = QtGui.QFont()
        font12.setFamily(u"MS Shell Dlg 2")
        font12.setPointSize(14)
        font12.setBold(True)
        font12.setWeight(75)
        
        font13 = QtGui.QFont()
        font13.setFamily(u"MS Shell Dlg 2")
        font13.setPointSize(10)
        font13.setBold(True)
        font13.setWeight(75)
        
        self.item_label_3.setFont(font3)
        self.item_label_3.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_3.setWordWrap(True)
        self.RollsQuantity = QtWidgets.QTextEdit(self.RollsContainer)
        self.RollsQuantity.setObjectName(u"RollsQuantity")
        self.RollsQuantity.setGeometry(QtCore.QRect(520, 110, 111, 31))
        self.RollsQuantity.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_4 = QtWidgets.QLabel(self.RollsContainer)
        self.item_label_4.setObjectName(u"item_label_4")
        self.item_label_4.setGeometry(QtCore.QRect(330, 110, 101, 31))
        self.item_label_4.setFont(font3)
        self.item_label_4.setTextFormat(QtCore.Qt.RichText)
        self.item_label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_4.setWordWrap(True)
        self.RollsRate = QtWidgets.QTextEdit(self.RollsContainer) 
        self.RollsRate.setObjectName(u"RollsRate")
        self.RollsRate.setGeometry(QtCore.QRect(110, 160, 181, 31))
        self.RollsRate.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_5 = QtWidgets.QLabel(self.RollsContainer)
        self.item_label_5.setObjectName(u"item_label_5")
        self.item_label_5.setGeometry(QtCore.QRect(30, 160, 61, 31))
        self.item_label_5.setFont(font3)
        self.item_label_5.setTextFormat(QtCore.Qt.RichText)
        self.item_label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_5.setWordWrap(True)
        self.Add_Rolls = QtWidgets.QPushButton(self.RollsContainer)
        self.Add_Rolls.setObjectName(u"Add_Rolls")
        self.Add_Rolls.setGeometry(QtCore.QRect(300, 220, 111, 51))
        self.Add_Rolls.setFont(font2)
        self.Add_Rolls.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
         
        self.groupBox3 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox3.setObjectName(u"groupBox3")
        self.groupBox3.setGeometry(QtCore.QRect(758, 97, 701, 91))
        self.groupBox3.setFont(font2)
        self.groupBox3.setStyleSheet(u"background-color: rgb(126, 255, 247);\n""color:rgb(0,0,81) ;")
        self.prev_balance = QtWidgets.QLabel(self.groupBox3)
        self.prev_balance.setObjectName(u"prev_balance")
        self.prev_balance.setGeometry(QtCore.QRect(15,10,171,31))
        self.prev_balance.setFont(font3)
        self.prev_balance.setTextFormat(QtCore.Qt.PlainText)
        self.prev_balance.setAlignment(QtCore.Qt.AlignCenter)
        self.prev_balance.setWordWrap(True)
        self.prevbalance = QtWidgets.QTextEdit(self.groupBox3)
        self.prevbalance.setObjectName(u"prevbalance")
        self.prevbalance.setGeometry(QtCore.QRect(240, 10,91, 31))
        self.prevbalance.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.prevbalance.setEnabled(False)
        self.prevbalance.setText(str(self.clientrecord[3]))
        
        self.m_ledger = QtWidgets.QTextEdit(self.groupBox3)
        self.m_ledger.setObjectName(u"m_ledger")
        self.m_ledger.setGeometry(QtCore.QRect(520, 10, 170, 31))
        self.m_ledger.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.m_ledger.setText(str(self.clientrecord[-2]))
        self.manualledger = QtWidgets.QLabel(self.groupBox3)
        self.manualledger.setObjectName(u"balance")
        self.manualledger.setGeometry(QtCore.QRect(400,10, 100, 31))
        self.manualledger.setFont(font3)
        self.manualledger.setText("Balance:")
        
        self.prev_rcp = QtWidgets.QLabel(self.groupBox3)
        self.prev_rcp.setObjectName(u"prev_rcp")
        self.prev_rcp.setGeometry(QtCore.QRect(0, 50, 211, 31))
        self.prev_rcp.setFont(font3)
        self.prev_rcp.setTextFormat(QtCore.Qt.PlainText)
        self.prev_rcp.setAlignment(QtCore.Qt.AlignCenter)
        self.prev_rcp.setWordWrap(True)
        self.prev_rcp.setText("Previous Receipt:")

        self.prevrcp = QtWidgets.QTextEdit(self.groupBox3)
        self.prevrcp.setObjectName(u"prevrcp")
        self.prevrcp.setGeometry(QtCore.QRect(240, 50,91,31))
        self.prevrcp.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.prevrcp.setEnabled(False)
        self.prevrcp.setText(str(self.clientrecord[-1]))    
        self.prevrcp.setEnabled(False)

         #[CLIENT_ID,CLIENT_NAME,CONTACT_NO,BALANCE,LEDGER]
            
            
        self.item_label_19 = QtWidgets.QLabel(self.RollsContainer)
        self.item_label_19.setObjectName(u"item_label_19")
        self.item_label_19.setGeometry(QtCore.QRect(330, 30, 171, 61))
        self.item_label_19.setFont(font3)
        self.item_label_19.setTextFormat(QtCore.Qt.RichText)
        self.item_label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_19.setWordWrap(True)
        self.item_label_7 = QtWidgets.QLabel(self.RollsContainer)
        self.item_label_7.setObjectName(u"item_label_7")
        self.item_label_7.setGeometry(QtCore.QRect(30, 90, 51, 51))
        self.item_label_7.setFont(font3)
        self.item_label_7.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_7.setWordWrap(True)
        self.RollsSizeComboBox = QtWidgets.QComboBox(self.RollsContainer)
        self.RollsSizeComboBox.setObjectName(u"RollsSizeComboBox")
        self.RollsSizeComboBox.setGeometry(QtCore.QRect(100, 100, 191, 31))
        self.RollsSizeComboBox.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.RollsSizeComboBox.addItem("")
        self.QuantityStockRolls = QtWidgets.QTextEdit(self.RollsContainer)
        self.QuantityStockRolls.setObjectName(u"qtysr")
        self.QuantityStockRolls.setEnabled(False)
        self.QuantityStockRolls.setGeometry(QtCore.QRect(520, 40, 111, 31))
        self.QuantityStockRolls.setStyleSheet(u"background-color: rgb(255, 255, 255);")

        
        self.ReelsContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.ReelsContainer.setObjectName(u"ReelsContainer")
        self.ReelsContainer.setGeometry(QtCore.QRect(752, 240, 661, 301))
        self.ReelsContainer.setFont(font2)
        self.ReelsContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.item_label_11 = QtWidgets.QLabel(self.ReelsContainer)
        self.item_label_11.setObjectName(u"item_label_11")
        self.item_label_11.setGeometry(QtCore.QRect(20, 20, 61, 51))
        self.item_label_11.setFont(font1)
        self.item_label_11.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_11.setWordWrap(True)
        
        self.ReelsRate = QtWidgets.QTextEdit(self.ReelsContainer)
        self.ReelsRate.setObjectName(u"ReelsRate")
        self.ReelsRate.setGeometry(QtCore.QRect(560, 30, 81, 31))
        self.ReelsRate.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_26 = QtWidgets.QLabel(self.ReelsContainer)
        self.item_label_26.setObjectName(u"item_label_26")
        self.item_label_26.setGeometry(QtCore.QRect(370, 30, 61, 31))
        self.item_label_26.setFont(font1)
        self.item_label_26.setTextFormat(QtCore.Qt.RichText)
        self.item_label_26.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_26.setWordWrap(True)
        self.item_label_27 = QtWidgets.QLabel(self.ReelsContainer)
        self.item_label_27.setObjectName(u"item_label_27")
        self.item_label_27.setGeometry(QtCore.QRect(370, 90, 181, 41))
        self.item_label_27.setFont(font1)
        self.item_label_27.setTextFormat(QtCore.Qt.RichText)
        self.item_label_27.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_27.setWordWrap(True)
        self.item_label_28 = QtWidgets.QLabel(self.ReelsContainer)
        self.item_label_28.setObjectName(u"item_label_28")
        self.item_label_28.setGeometry(QtCore.QRect(20, 80, 51, 51))
        self.item_label_28.setFont(font1)
        self.item_label_28.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_28.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_28.setWordWrap(True)
        self.ReelsPaperTypeComboBox = QtWidgets.QComboBox(self.ReelsContainer)
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.addItem("")
        self.ReelsPaperTypeComboBox.setObjectName(u"ReelsPaperTypeComboBox")
        self.ReelsPaperTypeComboBox.setGeometry(QtCore.QRect(150,30, 191, 31))
        self.ReelsPaperTypeComboBox.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.QuantityStockReels = QtWidgets.QTextEdit(self.ReelsContainer)
        self.QuantityStockReels.setObjectName(u"QuantityStockReels")
        self.QuantityStockReels.setEnabled(False)
        self.QuantityStockReels.setGeometry(QtCore.QRect(560, 90, 81, 31))
        self.QuantityStockReels.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.ReelsSizeComboBox = QtWidgets.QComboBox(self.ReelsContainer)
         
        self.ReelsSizeComboBox.setObjectName(u"ReelsSizeComboBox")
        self.ReelsSizeComboBox.setGeometry(QtCore.QRect(150, 90, 191, 31))
        self.ReelsSizeComboBox.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.ReelsSizeComboBox.addItem("")
        self.item_label_30 = QtWidgets.QLabel(self.ReelsContainer)
        self.item_label_30.setObjectName(u"item_label_30")
        self.item_label_30.setGeometry(QtCore.QRect(20, 140, 121, 41))
        self.item_label_30.setFont(font1)
        self.item_label_30.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_30.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_30.setWordWrap(True)
        self.ReelsWeight= QtWidgets.QComboBox(self.ReelsContainer)
        self.ReelsWeight.addItem("")
        self.ReelsWeight.setObjectName(u"ReelsWeightComboBox")
        self.ReelsWeight.setGeometry(QtCore.QRect(150, 150, 191, 31))
        self.ReelsWeight.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.Add_Reels = QtWidgets.QPushButton(self.ReelsContainer)
        self.Add_Reels.setObjectName(u"Add_Reels")
        self.Add_Reels.setGeometry(QtCore.QRect(280, 230, 111, 51))
        self.Add_Reels.setFont(font2)
        self.Add_Reels.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.TotayContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.TotayContainer.setObjectName(u"TotayContainer")
        self.TotayContainer.setGeometry(QtCore.QRect(752, 240, 581, 401))
        self.TotayContainer.setFont(font2)
        self.TotayContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
         
        self.item_label_18 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_18.setObjectName(u"item_label_18")
        self.item_label_18.setGeometry(QtCore.QRect(280, 50, 61, 31))
        self.item_label_18.setFont(font1)
        self.item_label_18.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_18.setWordWrap(True)
        self.TotaSizeTextField = QtWidgets.QComboBox(self.TotayContainer)
        self.TotaSizeTextField.setObjectName(u"TotaSizeTextField")
        self.TotaSizeTextField.setGeometry(QtCore.QRect(400,50, 161, 31))
        self.TotaSizeTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.TotaSizeTextField.addItem("")
        
        self.item_label_17= QtWidgets.QLabel(self.TotayContainer)
        self.item_label_17.setObjectName(u"item_label_17")
        self.item_label_17.setGeometry(QtCore.QRect(10, 40, 61, 51))
        self.item_label_17.setFont(font1)
        self.item_label_17.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_17.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_17.setWordWrap(True)
        
        self.totaTypeComboBox_2 = QtWidgets.QComboBox(self.TotayContainer)
        self.totaTypeComboBox_2.setObjectName(u"TotatypeField")
        self.totaTypeComboBox_2.setGeometry(QtCore.QRect(130,50, 131, 31))
        self.totaTypeComboBox_2.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")
        self.totaTypeComboBox_2.addItem("")

        #self.totaTypeComboBox_2.setText("0")
        self.item_label_17.setText("Type:")
        
        
        self.totayquantityfield = QtWidgets.QTextEdit(self.TotayContainer)
        self.totayquantityfield.setObjectName(u"toooot")
        self.totayquantityfield.setGeometry(QtCore.QRect( 470,120,91,31))
        self.totayquantityfield.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.totayquantityfield.setEnabled(False)
        self.item_label_41 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_41.setObjectName(u"itm")
        self.item_label_41.setGeometry(QtCore.QRect(280,120, 181, 31))
        self.item_label_41.setFont(font1)
        self.item_label_41.setTextFormat(QtCore.Qt.RichText)
        self.item_label_41.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_41.setWordWrap(True)
        self.totayquantityfield.setText("0")
        self.item_label_41.setText("Quantity in Stock:")
        
        self.totaquantity = QtWidgets.QTextEdit(self.TotayContainer)
        self.totaquantity.setObjectName(u"TotaDetails")
        self.totaquantity.setGeometry(QtCore.QRect( 130,120,131,31))
        self.totaquantity.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_15 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_15.setObjectName(u"item_label_15")
        self.item_label_15.setGeometry(QtCore.QRect(10,120, 101, 31))
        self.item_label_15.setFont(font1)
        self.item_label_15.setTextFormat(QtCore.Qt.RichText)
        self.item_label_15.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_15.setWordWrap(True)
        self.item_label_15.setText("Quantity:")
        self.totaquantity.setText("1")
        self.totaquantity.setEnabled(False) 
        
        
        
        
        
        self.TotaWeightTextField_2 = QtWidgets.QTextEdit(self.TotayContainer)
        self.TotaWeightTextField_2.setObjectName(u"packetgram_m")
        self.TotaWeightTextField_2.setGeometry(QtCore.QRect( 470,200,91,31))
        self.TotaWeightTextField_2.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_23 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_23.setObjectName(u"label")
        self.item_label_23.setGeometry(QtCore.QRect(290,190,161,61))
        self.item_label_23.setFont(font13)
        self.item_label_23.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_23.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_23.setWordWrap(True)
        self.item_label_23.setText("Add Weight Manually:") 
        self.item_label_23.setText("") 
        self.item_label_40 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_40.setObjectName(u"s2l")
        self.item_label_40.setGeometry(QtCore.QRect(290, 190,21, 51))
        self.item_label_40.setFont(font12)
        self.item_label_40.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_40.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_40.setWordWrap(True)
        
        
        self.Add_Totay = QtWidgets.QPushButton(self.TotayContainer)
        self.Add_Totay.setObjectName(u"Add_Totay")
        self.Add_Totay.setGeometry(QtCore.QRect(260,330,111,51 ))
        self.Add_Totay.setFont(font2)
        self.Add_Totay.setText('Add')
        self.Add_Totay.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        
        self.TotaRateTextField = QtWidgets.QTextEdit(self.TotayContainer)
        self.TotaRateTextField.setObjectName(u"TotaRateTextField")
        self.TotaRateTextField.setGeometry(QtCore.QRect(130,270,141,31))
        self.TotaRateTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_34 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_34.setObjectName(u"item_label_34")
        self.item_label_34.setGeometry(QtCore.QRect(10,270,61,31))
        self.item_label_34.setFont(font1)
        self.item_label_34.setTextFormat(QtCore.Qt.RichText)
        self.item_label_34.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_34.setWordWrap(True)
        self.item_label_22 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_22.setObjectName(u"item_label_22")
        self.item_label_22.setGeometry(QtCore.QRect(10,190,91,51))
        self.item_label_22.setFont(font1)
        self.item_label_22.setTextFormat(QtCore.Qt.RichText)
        self.item_label_22.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_22.setWordWrap(True)
        self.TotaWeightTextField = QtWidgets.QComboBox(self.TotayContainer)
        self.TotaWeightTextField.setObjectName(u"TotaWeightTextField")
        self.TotaWeightTextField.setGeometry(QtCore.QRect(130,200,141,31))
        self.TotaWeightTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.NaliContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.NaliContainer.setObjectName(u"NaliContainer")
        self.NaliContainer.setGeometry(QtCore.QRect(752, 240, 381, 221))
        self.NaliContainer.setFont(font2)
        self.NaliContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.Add_Nali = QtWidgets.QPushButton(self.NaliContainer)
        self.Add_Nali.setObjectName(u"Add_Nali")
        self.Add_Nali.setGeometry(QtCore.QRect(140, 150, 111, 51))
        self.Add_Nali.setFont(font2)
        self.Add_Nali.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.NaliRateTextField = QtWidgets.QTextEdit(self.NaliContainer)
        self.NaliRateTextField.setObjectName(u"NaliRateTextField")
        self.NaliRateTextField.setGeometry(QtCore.QRect(210, 90, 111, 31))
        self.NaliRateTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.NaliRate = QtWidgets.QLabel(self.NaliContainer)
        self.NaliRate.setObjectName(u"NaliRate")
        self.NaliRate.setGeometry(QtCore.QRect(100, 90, 61, 31))
        self.NaliRate.setFont(font1)
        self.NaliRate.setTextFormat(QtCore.Qt.RichText)
        self.NaliRate.setAlignment(QtCore.Qt.AlignCenter)
        self.NaliRate.setWordWrap(True)
        self.NaliWeightTextField = QtWidgets.QTextEdit(self.NaliContainer)
        self.NaliWeightTextField.setObjectName(u"NaliWeightTextField")
        self.NaliWeightTextField.setGeometry(QtCore.QRect(210, 30, 111, 31))
        self.NaliWeightTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.NaliWeight = QtWidgets.QLabel(self.NaliContainer)
        self.NaliWeight.setObjectName(u"NaliWeight")
        self.NaliWeight.setGeometry(QtCore.QRect(100, 20, 81, 51))
        self.NaliWeight.setFont(font1)
        self.NaliWeight.setTextFormat(QtCore.Qt.RichText)
        self.NaliWeight.setAlignment(QtCore.Qt.AlignCenter)
        self.NaliWeight.setWordWrap(True)
        self.JuttaContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.JuttaContainer.setObjectName(u"JuttaContainer")
        self.JuttaContainer.setGeometry(QtCore.QRect(752, 240, 401, 259))
        self.JuttaContainer.setFont(font2)
        self.JuttaContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.Add_Jutta = QtWidgets.QPushButton(self.JuttaContainer)
        self.Add_Jutta.setObjectName(u"Add_Jutta")
        self.Add_Jutta.setGeometry(QtCore.QRect(150,180, 111, 51))
        self.Add_Jutta.setFont(font2)
        self.Add_Jutta.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        
        self.JuttaRateTextField = QtWidgets.QTextEdit(self.JuttaContainer)
        self.JuttaRateTextField.setObjectName(u"JuttaRateTextField")
        self.JuttaRateTextField.setGeometry(QtCore.QRect(240, 120, 111, 31))
        self.JuttaRateTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_38 = QtWidgets.QLabel(self.JuttaContainer)
        self.item_label_38.setObjectName(u"item_label_38")
        self.item_label_38.setGeometry(QtCore.QRect(80, 120, 61, 31))
        self.item_label_38.setFont(font1)
        self.item_label_38.setTextFormat(QtCore.Qt.RichText)
        self.item_label_38.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_38.setWordWrap(True)
        self.JuttaWeightTextField = QtWidgets.QTextEdit(self.JuttaContainer)
        self.JuttaWeightTextField.setObjectName(u"JuttaWeightTextField")
        self.JuttaWeightTextField.setGeometry(QtCore.QRect(240, 60, 111, 31))
        self.JuttaWeightTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_37 = QtWidgets.QLabel(self.JuttaContainer)
        self.item_label_37.setObjectName(u"item_label_37")
        self.item_label_37.setGeometry(QtCore.QRect(80, 50, 131, 51))
        self.item_label_37.setFont(font1)
        self.item_label_37.setTextFormat(QtCore.Qt.RichText)
        self.item_label_37.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_37.setWordWrap(True)
        
        
        self.RaddiContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.RaddiContainer.setObjectName(u"RaddiContainer")
        self.RaddiContainer.setGeometry(QtCore.QRect(752, 240, 411, 261))
        self.RaddiContainer.setFont(font2)
        self.RaddiContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.Add_Raddi = QtWidgets.QPushButton(self.RaddiContainer)
        self.Add_Raddi.setObjectName(u"Add_Raddi")
        self.Add_Raddi.setGeometry(QtCore.QRect(180, 190, 111, 51))
        self.Add_Raddi.setFont(font2)
        self.Add_Raddi.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.RaddiRateTextField = QtWidgets.QTextEdit(self.RaddiContainer)
        self.RaddiRateTextField.setObjectName(u"RaddiRateTextField")
        self.RaddiRateTextField.setGeometry(QtCore.QRect(190, 120, 121, 31))
        self.RaddiRateTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.RaddiRate = QtWidgets.QLabel(self.RaddiContainer)
        self.RaddiRate.setObjectName(u"RaddiRate")
        self.RaddiRate.setGeometry(QtCore.QRect(90, 120, 61, 31))
        self.RaddiRate.setFont(font1)
        self.RaddiRate.setTextFormat(QtCore.Qt.RichText)
        self.RaddiRate.setAlignment(QtCore.Qt.AlignCenter)
        self.RaddiRate.setWordWrap(True)
        self.RaddiWeightTextField = QtWidgets.QTextEdit(self.RaddiContainer)
        self.RaddiWeightTextField.setObjectName(u"RaddiWeightTextField")
        self.RaddiWeightTextField.setGeometry(QtCore.QRect(190, 50, 121, 31))
        self.RaddiWeightTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.RaddiWeight = QtWidgets.QLabel(self.RaddiContainer)
        self.RaddiWeight.setObjectName(u"RaddiWeight")
        self.RaddiWeight.setGeometry(QtCore.QRect(90, 40, 81, 51))
        self.RaddiWeight.setFont(font1)
        self.RaddiWeight.setTextFormat(QtCore.Qt.RichText)
        self.RaddiWeight.setAlignment(QtCore.Qt.AlignCenter)
        self.RaddiWeight.setWordWrap(True)
        self.PacketsContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.PacketsContainer.setGeometry(QtCore.QRect(752, 240, 701,501))
 
        self.PacketsContainer.setFont(font2)
        self.PacketsContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        
        self.item_label2 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label2.setObjectName(u"item_label2")
        self.item_label2.setGeometry(QtCore.QRect(30, 40, 81, 51))
        self.item_label2.setFont(font1)
        self.item_label2.setTextFormat(QtCore.Qt.PlainText)
        self.item_label2.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label2.setWordWrap(True)
        self.Packetslength = QtWidgets.QTextEdit(self.PacketsContainer)
        self.Packetslength.setObjectName(u"Packetslength")
        self.Packetslength.setGeometry(QtCore.QRect(130,50, 191, 31))
        self.Packetslength.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.PacketsContainer.setInputMethodHints(QtCore.Qt.ImhNone)
        self.PacketsContainer.setFlat(False)
        self.PacketsNoOfPackets = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsNoOfPackets.setObjectName(u"PacketsNoOfPackets")
        self.PacketsNoOfPackets.setGeometry(QtCore.QRect(550, 130, 101, 31))
        self.PacketsNoOfPackets.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_14 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_14.setObjectName(u"item_label_14")
        self.item_label_14.setGeometry(QtCore.QRect(340, 110, 141, 71))
        self.item_label_14.setFont(font1)
        self.item_label_14.setTextFormat(QtCore.Qt.RichText)
        self.item_label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_14.setWordWrap(True)
        self.item_label_13 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_13.setObjectName(u"item_label_13")
        self.item_label_13.setGeometry(QtCore.QRect(20,120, 81, 51))
        self.item_label_13.setFont(font1)
        self.item_label_13.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_13.setWordWrap(True)
        
        
        self.PacketsRate = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsRate.setObjectName(u"PacketsRate")
        self.PacketsRate.setGeometry(QtCore.QRect(210,200, 111, 31))
        self.PacketsRate.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.PacketsPaperTypeComboBox = QtWidgets.QComboBox(self.PacketsContainer)
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.setObjectName(u"PacketsPaperTypeComboBox")
        self.PacketsPaperTypeComboBox.setGeometry(QtCore.QRect(130,130, 191, 31))
        self.PacketsPaperTypeComboBox.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_31 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_31.setObjectName(u"item_label_31")
        self.item_label_31.setGeometry(QtCore.QRect(30, 200, 71, 31))
        self.item_label_31.setFont(font1)
        self.item_label_31.setTextFormat(QtCore.Qt.RichText)
        self.item_label_31.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_31.setWordWrap(True)
        self.Add_Packets = QtWidgets.QPushButton(self.PacketsContainer)
        self.Add_Packets.setObjectName(u"Add_Packets")
        self.Add_Packets.setGeometry(QtCore.QRect(300, 420, 111, 51))
        self.Add_Packets.setFont(font2)
        self.Add_Packets.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.PacketsGrammage = QtWidgets.QComboBox(self.PacketsContainer)
        self.PacketsGrammage.setObjectName(u"PacketsGrammage")
        self.PacketsGrammage.setGeometry(QtCore.QRect(460, 360, 191, 31))
        self.PacketsGrammage.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_33 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_33.setObjectName(u"item_label_33")
        self.item_label_33.setGeometry(QtCore.QRect(360,350, 81, 41))
        self.item_label_33.setFont(font1)
        self.item_label_33.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_33.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_33.setWordWrap(True)
        
        self.packetgram_m = QtWidgets.QTextEdit(self.PacketsContainer)
        self.packetgram_m.setObjectName(u"packetgram_m")
        self.packetgram_m.setGeometry(QtCore.QRect(550,200, 101, 31))
        self.packetgram_m.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.packetgram_m_label = QtWidgets.QLabel(self.PacketsContainer)
        self.packetgram_m_label.setObjectName(u"packetgram_m_label")
        self.packetgram_m_label.setGeometry(QtCore.QRect(350,190,141, 51))
        self.packetgram_m_label.setFont(font1)
        self.packetgram_m_label.setTextFormat(QtCore.Qt.PlainText)
        self.packetgram_m_label.setAlignment(QtCore.Qt.AlignCenter)
        self.packetgram_m_label.setWordWrap(True)
        self.packetgram_m_label.setText("Grammage:") 
        self.packetgram_m.setText("") 
        '''self.item_label_36 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_36.setObjectName(u"sl")
        self.item_label_36.setGeometry(QtCore.QRect(370, 90,41, 41))
        self.item_label_36.setFont(font12)
        self.item_label_36.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_36.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_36.setWordWrap(True) '''
        
         
        
        self.PacketsTotaDetails = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsTotaDetails.setObjectName(u"PacketsTotaDetails")
        self.PacketsTotaDetails.setGeometry(QtCore.QRect(220,360, 101, 31))
        self.PacketsTotaDetails.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.PacketsTotaDetails.setEnabled(False)
        
        self.item_label_16 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_16.setObjectName(u"item_label_16")
        self.item_label_16.setGeometry(QtCore.QRect(350, 40, 81, 51))
        self.item_label_16.setFont(font1)
        self.item_label_16.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_16.setWordWrap(True)
        self.item_label_20 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_20.setObjectName(u"item_label_20")
        self.item_label_20.setGeometry(QtCore.QRect(10,350, 151, 61))
        self.item_label_20.setFont(font1)
        self.item_label_20.setTextFormat(QtCore.Qt.RichText)
        self.item_label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_20.setWordWrap(True)
        self.PacketsWidth = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsWidth.setObjectName(u"PacketsWidth")
        self.PacketsWidth.setGeometry(QtCore.QRect(550, 50, 101, 31))
        self.PacketsWidth.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_36 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_36.setObjectName(u"item_36")
        self.item_36.setGeometry(QtCore.QRect(20,270, 91, 41))
        self.item_36.setFont(font1)
        self.item_36.setTextFormat(QtCore.Qt.PlainText)
        self.item_36.setAlignment(QtCore.Qt.AlignCenter)
        self.item_36.setWordWrap(True)
        self.items = QtWidgets.QComboBox(self.PacketsContainer)
        self.items.setObjectName(u"items")
        self.items.setGeometry(QtCore.QRect(130,280,191,31))
        self.items.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.items.addItem("")
        self.items.addItem("")
        self.items.addItem("")
        self.packetsize = QtWidgets.QComboBox(self.PacketsContainer)
        self.packetsize.setObjectName(u"packetsize")
        self.packetsize.setGeometry(QtCore.QRect(460,280,191,31))
        self.packetsize.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.packetsize.addItem("")
        
        self.item_label_35 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_35.setObjectName(u"item_label_35")
        self.item_label_35.setGeometry(QtCore.QRect(360,270,81,51))
        self.item_label_35.setFont(font1)
        self.item_label_35.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_35.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_35.setWordWrap(True)
        
        '''
         
        self.PacketsContainer = QtWidgets.QGroupBox(self.centralwidget)
        self.PacketsContainer.setGeometry(QtCore.QRect(1000,300, 701, 501))
        self.PacketsContainer.setFont(font2)
        self.PacketsContainer.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.PacketsContainer.setInputMethodHints(QtCore.Qt.ImhNone)
        self.PacketsContainer.setFlat(False)
        self.PacketsNoOfPackets = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsNoOfPackets.setObjectName(u"PacketsNoOfPackets")
        self.PacketsNoOfPackets.setGeometry(QtCore.QRect(550, 130, 101, 31))
        self.PacketsNoOfPackets.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_14 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_14.setObjectName(u"item_label_14")
        self.item_label_14.setGeometry(QtCore.QRect(340, 110, 141, 71))
        self.item_label_14.setFont(font1)
        self.item_label_14.setTextFormat(QtCore.Qt.RichText)
        self.item_label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_14.setWordWrap(True)
        self.item_label_13 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_13.setObjectName(u"item_label_13")
        self.item_label_13.setGeometry(QtCore.QRect(30, 200, 71, 31))
        self.item_label_13.setFont(font1)
        self.item_label_13.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_13.setWordWrap(True)
        self.item_label_13.setText("Type:")
        
        self.PacketsRate = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsRate.setObjectName(u"PacketsRate")
        self.PacketsRate.setGeometry(QtCore.QRect(240,240 , 111, 31))
        self.PacketsRate.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.PacketsPaperTypeComboBox = QtWidgets.QComboBox(self.PacketsContainer)
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.addItem("")
        self.PacketsPaperTypeComboBox.setObjectName(u"PacketsPaperTypeComboBox")
        self.PacketsPaperTypeComboBox.setGeometry(QtCore.QRect(130,130, 191, 31))
        self.PacketsPaperTypeComboBox.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_31 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_31.setObjectName(u"item_label_31")
        self.item_label_31.setGeometry(QtCore.QRect(30, 200, 71, 31))
        self.item_label_31.setFont(font1)
        self.item_label_31.setTextFormat(QtCore.Qt.RichText)
        self.item_label_31.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_31.setWordWrap(True)
        self.Add_Packets = QtWidgets.QPushButton(self.PacketsContainer)
        self.Add_Packets.setObjectName(u"Add_Packets")
        self.Add_Packets.setGeometry(QtCore.QRect(300, 420, 111, 51))
        self.Add_Packets.setFont(font2)
        self.Add_Packets.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.PacketsGrammage = QtWidgets.QComboBox(self.PacketsContainer)
        self.PacketsGrammage.setObjectName(u"PacketsGrammage")
        self.PacketsGrammage.setGeometry(QtCore.QRect(460,280,191,31))
        self.PacketsGrammage.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.item_label_33 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_33.setObjectName(u"item_label_33")
        self.item_label_33.setGeometry(QtCore.QRect(360,270,81,51))
        self.item_label_33.setFont(font1)
        self.item_label_33.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_33.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_33.setWordWrap(True)
        
        self.packetgram_m = QtWidgets.QTextEdit(self.PacketsContainer)
        self.packetgram_m.setObjectName(u"packetgram_m")
        self.packetgram_m.setGeometry(QtCore.QRect(550,200, 101, 31))
        self.packetgram_m.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.packetgram_m_label = QtWidgets.QLabel(self.PacketsContainer)
        self.packetgram_m_label.setObjectName(u"packetgram_m_label")
        self.packetgram_m_label.setGeometry(QtCore.QRect(350,190,141, 51))
        self.packetgram_m_label.setFont(font13)
        self.packetgram_m_label.setTextFormat(QtCore.Qt.PlainText)
        self.packetgram_m_label.setAlignment(QtCore.Qt.AlignCenter)
        self.packetgram_m_label.setWordWrap(True)
        self.packetgram_m_label.setText("Grammage:") 
        self.packetgram_m.setText("") 
        
        self.item_label_36 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_36.setObjectName(u"sl")
        self.item_label_36.setGeometry(QtCore.QRect(20,270,91,41))
        self.item_label_36.setFont(font12)
        self.item_label_36.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_36.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_36.setWordWrap(True)
        self.item_label_36.setText("Tota/Reels:") 
        
        
        
        self.PacketsTotaDetails = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsTotaDetails.setObjectName(u"PacketsTotaDetails")
        self.PacketsTotaDetails.setGeometry(QtCore.QRect(220,360, 101, 31))
        self.PacketsTotaDetails.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.PacketsTotaDetails.setEnabled(False)
        self.item_label_16 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_16.setObjectName(u"item_label_16")
        self.item_label_16.setGeometry(QtCore.QRect(30, 200, 68, 24))
        self.item_label_16.setFont(font1)
        self.item_label_16.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_16.setWordWrap(True)
        self.item_label_20 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_20.setObjectName(u"item_label_20")
        self.item_label_20.setGeometry(QtCore.QRect(10,350, 151, 61))
        self.item_label_20.setFont(font1)
        self.item_label_20.setTextFormat(QtCore.Qt.RichText)
        self.item_label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_20.setWordWrap(True)
        self.PacketsWidth = QtWidgets.QTextEdit(self.PacketsContainer)
        self.PacketsWidth.setObjectName(u"PacketsWidth")
        self.PacketsWidth.setGeometry(QtCore.QRect(160, 200, 191, 31))
        self.PacketsWidth.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        
        self.packetsize = QtWidgets.QComboBox(self.PacketsContainer)
        self.packetsize.setObjectName(u"packetsize")
        self.packetsize.setGeometry(QtCore.QRect(460, 360, 191, 31))
        self.packetsize.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.packetsize.addItem("")
        self.items = QtWidgets.QComboBox(self.PacketsContainer)
        self.items.setObjectName(u"items")
        self.items.setGeometry(QtCore.QRect(130,280,191,31))
        self.items.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.items.addItem("")
        self.items.addItem("")
        self.items.addItem("")
        self.item_36 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_36.setObjectName(u"item_36")
        self.item_36.setGeometry(QtCore.QRect(20,270, 91, 41))
        self.item_36.setFont(font1)
        self.item_36.setTextFormat(QtCore.Qt.PlainText)
        self.item_36.setAlignment(QtCore.Qt.AlignCenter)
        self.item_36.setWordWrap(True)
        
        self.item_label_35 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label_35.setObjectName(u"item_label_35")
        self.item_label_35.setGeometry(QtCore.QRect(360,350, 51, 41))
        self.item_label_35.setFont(font1)
        self.item_label_35.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_35.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_35.setWordWrap(True)
        
        self.item_label2 = QtWidgets.QLabel(self.PacketsContainer)
        self.item_label2.setObjectName(u"item_label2")
        self.item_label2.setGeometry(QtCore.QRect(30, 40, 81, 51))
        self.item_label2.setFont(font1)
        self.item_label2.setTextFormat(QtCore.Qt.PlainText)
        self.item_label2.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label2.setWordWrap(True)
        self.Packetslength = QtWidgets.QTextEdit(self.PacketsContainer)
        self.Packetslength.setObjectName(u"Packetslength")
        self.Packetslength.setGeometry(QtCore.QRect(130,150, 191, 31))
        self.Packetslength.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        ''' 
        self.CheckOutButton = QtWidgets.QPushButton(self.centralwidget)
        self.CheckOutButton.setGeometry(QtCore.QRect(512, 700, 231, 71))
        self.CheckOutButton.setObjectName("CheckOutButton")
        self.CheckOutButton.setFont(font2)
        self.CheckOutButton.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.textBrowser_3 = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser_3.setObjectName(u"textBrowser_3")
        self.textBrowser_3.setEnabled(False)
        self.textBrowser_3.setGeometry(QtCore.QRect(10, 960, 1901, 31))
        self.textBrowser_3.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"gridline-color: rgb(0, 0, 127);\n"
"color:rgb(255,255, 255) ;")
         
        self.groupBox_2 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_2.setObjectName(u"groupBox_2")
        self.groupBox_2.setGeometry(QtCore.QRect(2, 612, 371, 241))
        self.groupBox_2.setFont(font2)
        self.groupBox_2.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.item_label_24 = QtWidgets.QLabel(self.groupBox_2)
        self.item_label_24.setObjectName(u"item_label_24")
        self.item_label_24.setGeometry(QtCore.QRect(10, 150, 151, 71))
        font4 = QtGui.QFont()
        font4.setFamily(u"MS Shell Dlg 2")
        font4.setBold(True)
        font4.setWeight(75)
        font4.setPointSize(12)
        self.item_label_24.setFont(font4)
        self.item_label_24.setTextFormat(QtCore.Qt.RichText)
        self.item_label_24.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_24.setWordWrap(True)
        self.totalTextField = QtWidgets.QTextEdit(self.groupBox_2)
        self.totalTextField.setObjectName(u"totalTextField")
        self.totalTextField.setGeometry(QtCore.QRect(170, 10, 161, 31))
        self.totalTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        self.item_label_8 = QtWidgets.QLabel(self.groupBox_2)
        self.item_label_8.setObjectName(u"item_label_8")
        self.item_label_8.setGeometry(QtCore.QRect(20, 0, 100, 61))
        self.item_label_8.setFont(font4)
        self.item_label_8.setTextFormat(QtCore.Qt.RichText)
        self.item_label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_8.setWordWrap(True)
        self.rate_packets_12 = QtWidgets.QLabel(self.groupBox_2)
        self.rate_packets_12.setObjectName(u"rate_packets_12")
        self.rate_packets_12.setGeometry(QtCore.QRect(20,70,91, 31))
        self.rate_packets_12.setFont(font4)
        self.rate_packets_12.setTextFormat(QtCore.Qt.RichText)
        self.rate_packets_12.setAlignment(QtCore.Qt.AlignCenter)
        self.rate_packets_12.setWordWrap(True)
        self.CreditTextField = QtWidgets.QTextEdit(self.groupBox_2)
        self.CreditTextField.setObjectName(u"CreditTextField")
        self.CreditTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        self.CreditTextField.setGeometry(QtCore.QRect(170, 120, 161, 31))
        self.CreditDetailsTextField = QtWidgets.QTextEdit(self.groupBox_2)
        self.CreditDetailsTextField.setObjectName(u"CreditDetailsTextField")
        self.CreditDetailsTextField.setGeometry(QtCore.QRect(170, 170, 161, 31))
        self.CreditDetailsTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        self.item_label_25 = QtWidgets.QLabel(self.groupBox_2)
        self.item_label_25.setObjectName(u"item_label_25")
        self.item_label_25.setGeometry(QtCore.QRect(10, 120, 111, 31))
        self.item_label_25.setFont(font4)
        self.item_label_25.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_25.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_25.setWordWrap(True)
        self.RentTextField = QtWidgets.QTextEdit(self.groupBox_2)
        self.RentTextField.setObjectName(u"RentTextField")
        self.RentTextField.setGeometry(QtCore.QRect(170, 70, 161, 31))
        self.RentTextField.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setObjectName(u"groupBox")
        self.groupBox.setGeometry(QtCore.QRect(2, 97, 741, 91))
        self.groupBox.setFont(font2)
        self.groupBox.setStyleSheet(u"background-color: rgb(126, 255, 247);\n"
"color:rgb(0,0,81) ;")
        self.rcp_no = QtWidgets.QTextEdit(self.groupBox)
        self.rcp_no.setObjectName(u"rcp_no")
        self.rcp_no.setEnabled(False)
        self.rcp_no.setGeometry(QtCore.QRect(150, 50, 111, 31))
        self.rcp_no.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        clientssheets=pd.read_excel('book.xlsx', index_col=None,sheet_name=None )
        toremove=['reels_stock_in_out','tota_stock_in_out','rolls_stock_in_out','client info','rolls_stock','reels_stock','totay','Fluting', 'Fluting_Bareek', 'L1', 'L1_Bareek', 'L2', 'L2_Bareek', 'Test_Liner', 'Test_Liner_Bareek', 'Box_Board_2_5_No', 'Box_Board_2_5_No_Bareek', 'Box_Board_3_No', 'Box_Board_3_No_Bareek', 'Local_Kraft', 'Local_Kraft_Bareek', 'Imported_Kraft', 'Imported_Kraft_Bareek', 'Super_Fluting', 'Super_Fluting_Bareek']
 
        sheets=[]
        for i in clientssheets.keys():
            if i not in toremove:
                sheets.append(i) 
                
        datas=pd.read_excel('book.xlsx' ,header=0, index_col=None,sheet_name=sheets, usecols=['RECIEPT_NUMBER'])

        rcp_df=pd.concat(datas[frame] for frame in datas.keys())
        w=rcp_df['RECIEPT_NUMBER'].unique()
        if(len(w)!=0):
                w[::-1].sort()
                self.rcp_no.setText(str(w[0]+1)) 

        else:
                 self.rcp_no.setText(str(1))
        
        self.clientName = QtWidgets.QTextEdit(self.groupBox)
        self.clientName.setObjectName(u"clientName")
        self.clientName.setGeometry(QtCore.QRect(510, 10, 211, 31))
        self.clientName.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        
        self.item_label_6 = QtWidgets.QLabel(self.groupBox)
        self.item_label_6.setObjectName(u"item_label_6")
        self.item_label_6.setGeometry(QtCore.QRect(10, 10, 71, 31))
        self.item_label_6.setFont(font1)
        self.item_label_6.setTextFormat(QtCore.Qt.RichText)
        self.item_label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_6.setWordWrap(True)
        self.item_label_2 = QtWidgets.QLabel(self.groupBox)
        self.item_label_2.setObjectName(u"item_label_2")
        self.item_label_2.setGeometry(QtCore.QRect(10, 50, 131, 31))
        self.item_label_2.setFont(font1)
        self.item_label_2.setTextFormat(QtCore.Qt.RichText)
        self.item_label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_2.setWordWrap(True)
        self.client_contact_label = QtWidgets.QLabel(self.groupBox)
        self.client_contact_label.setObjectName(u"client_contact_label")
        self.client_contact_label.setGeometry(QtCore.QRect(280, 60, 211, 16))
        self.client_contact_label.setFont(font1)
        
        self.contact = QtWidgets.QTextEdit(self.groupBox)
        self.contact.setObjectName(u"contact")
        self.contact.setGeometry(QtCore.QRect(510, 50, 211, 31))
        self.contact.setStyleSheet(u"background-color: rgb(255, 255, 255);")

        self.client_name_label = QtWidgets.QLabel(self.groupBox)
        self.client_name_label.setObjectName(u"client_name_label")
        self.client_name_label.setGeometry(QtCore.QRect(280, 10, 171, 41))
        self.client_name_label.setFont(font1)
        self.datetoday = QtWidgets.QTextEdit(self.groupBox)
        self.datetoday.setObjectName(u"datetoday")
        self.datetoday.setEnabled(False)
        self.datetoday.setGeometry(QtCore.QRect(150, 10, 111, 31)) 
        self.datetoday.setText(str(date.today().strftime("%d-%m-%y") ))
        self.datetoday.setStyleSheet(u"background-color: rgb(255, 255, 255);")
        self.DELETEROWS = QtWidgets.QPushButton(self.centralwidget)
        self.DELETEROWS.setGeometry(QtCore.QRect(512, 612, 231, 71))
        self.DELETEROWS.setObjectName("DELETE ENTRY")
        self.DELETEROWS.setText('Delete Entry')
        self.DELETEROWS.setFont(font2)
        self.DELETEROWS.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"color:rgb(255,255, 255) ;")
        self.textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser.setObjectName(u"textBrowser")
        self.textBrowser.setEnabled(False)
        self.textBrowser.setGeometry(QtCore.QRect(0, 10, 1460, 85))
        self.textBrowser.setStyleSheet(u"background-color:rgb(0, 0, 81) ;\n"
"gridline-color: rgb(0, 0, 127);\n"
"color:rgb(255,255, 255) ;")
        self.Add_Jutta.setText('Add')
        
        clientname=str(self.clientrecord[1])
        contactnum=str(self.clientrecord[2])
        self.contact.setText(str(contactnum))
        self.clientName.setText(str(clientname))
        
        self.contact.setEnabled(False)
        self.clientName.setEnabled(False)
        self.rcp_no.setEnabled(False)
        
        self.item_label_23 = QtWidgets.QLabel(self.TotayContainer)
        self.item_label_23.setObjectName(u"label")
        self.item_label_23.setGeometry(QtCore.QRect(290,190,161,61))
        self.item_label_23.setFont(font13)
        self.item_label_23.setTextFormat(QtCore.Qt.PlainText)
        self.item_label_23.setAlignment(QtCore.Qt.AlignCenter)
        self.item_label_23.setWordWrap(True)
        self.item_label_23.setText("Add Weight Manually:") 
        # cashbill table

        self.cashbillDetailstable = QtWidgets.QTableWidget(self.centralwidget)
        self.cashbillDetailstable.setGeometry(QtCore.QRect(2, 200, 741, 400))
        self.cashbillDetailstable.setObjectName("cashbillDetailstable")
        self.cashbillDetailstable.setRowCount(0)
        self.cashbillDetailstable.setColumnCount(2)
        self.cashbillDetailstable.setHorizontalHeaderLabels(('Item', 'Price')) 
        header = self.cashbillDetailstable.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)# set header text
        
          
        self.cashbillDetailstable.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.RollsContainer.raise_()
        self.itemListComboBox.raise_()
        self.item_label.raise_()
        self.client_name_label.raise_()
        self.item_label_2.raise_()
        self.CheckOutButton.raise_()
        self.CreditDetailsTextField.raise_()
        self.RentTextField.raise_()
        self.CreditTextField.raise_()
        self.rate_packets_12.raise_()
        self.item_label_24.raise_()
        self.item_label_25.raise_()
        self.clientName.raise_()
        self.rcp_no.raise_()
        self.datetoday.raise_()
        self.item_label_6.raise_()
        self.cashbillDetailstable.raise_()
        self.CreditTextField.setText('0')
        self.RentTextField.setText('0')
        self.totalTextField.setText('0')
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.itemListComboBox.activated.connect(self.AppearItem)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def updatedtext(self):
        totalrows = self.cashbillDetailstable.rowCount()
        #print('total rows: ', totalrows)
        updated = 0

        if totalrows != 0:
            for i in range(0, totalrows):
                updated += (float(self.cashbillDetailstable.item(i, 1).text()))

        self.totalTextField.setText(str(round(updated,2)))
    def reset_raddi_container(self):
        self.RaddiWeightTextField.setText('')     
        self.RaddiRateTextField.setText('')   
            
    def reset_rolls_container(self):
        self.RollsSizeComboBox.setCurrentIndex(0) 
        self.RollsPaperTypeComboBox.setCurrentIndex(0)
        self.RollsQuantity.setText( "0")
        self.QuantityStockRolls.setText("0")
        self.RollsRate.setText("")
            
    def reset_reels_container(self):
        self.QuantityStockReels.setText("0")
        self.ReelsSizeComboBox.setCurrentIndex(0)
        self.ReelsPaperTypeComboBox.setCurrentIndex(0)
        #self.ReelsQuantity.setText("0")
        self.ReelsWeight.setCurrentIndex(0)
        self.ReelsRate.setText("")
        
                
    def reset_packets_container(self):
        
        self.PacketsGrammage.setCurrentIndex(0)
        self.PacketsWidth.setText("")
        self.PacketsPaperTypeComboBox.setCurrentIndex(0)
        self.PacketsNoOfPackets.setText("")
        self.PacketsRate.setText("")
        self.packetsize.setCurrentIndex(0)
         # self.packetsize.
        self.PacketsTotaDetails.setText("0")
        self.packetgram_m.setText("") 
        self.Packetslength.setText("") 
         
        
        '''for index in range(1,len(self.packetsize)):
            self.packetsize.removeItem(index)
        for index in range(1,len(self.PacketsGrammage)):
            self.PacketsGrammage.removeItem(index)'''
             
    def reset_jutta_container(self):
        self.JuttaWeightTextField.setText("")
        self.JuttaRateTextField.setText("")
             
    def reset_totay_container(self):
        self.TotaWeightTextField.setCurrentIndex(0)
        self.TotaRateTextField.setText("")
        self.TotaSizeTextField.setCurrentIndex(0)
        self.TotaWeightTextField_2.setText("")
        self.totayquantityfield.setText("0")
        self.totaquantity.setText("1")
        self.totaquantity.setEnabled(False) 
        self.totaTypeComboBox_2.setCurrentIndex(0)
             
    def reset_nali_container(self):
        self.NaliWeightTextField.setText("")
        self.NaliRateTextField.setText("")
         
    def AppearItem(self):
        self.reset_reels_container()
        self.reset_packets_container() 
        self.reset_jutta_container()
        self.reset_totay_container() 
        self.reset_raddi_container()
        self.reset_nali_container() 
        self.reset_rolls_container() 
        if (self.itemListComboBox.currentText() == 'Rolls'):
            self.ReelsContainer.setVisible(False)
            self.PacketsContainer.setVisible(False)
            self.JuttaContainer.setVisible(False)
            self.TotayContainer.setVisible(False)
            self.RaddiContainer.setVisible(False)
            self.NaliContainer.setVisible(False)
            self.RollsContainer.setVisible(True)
            
            
            
            
            
            

        elif (self.itemListComboBox.currentText() == 'Reels'):
            self.RollsContainer.setVisible(False)
            self.PacketsContainer.setVisible(False)
            self.JuttaContainer.setVisible(False)
            self.TotayContainer.setVisible(False)
            self.RaddiContainer.setVisible(False)
            self.NaliContainer.setVisible(False)
            self.ReelsContainer.setVisible(True)

        elif (self.itemListComboBox.currentText() == 'Packets'):
            self.RollsContainer.setVisible(False)
            self.ReelsContainer.setVisible(False)
            self.JuttaContainer.setVisible(False)
            self.TotayContainer.setVisible(False)
            self.RaddiContainer.setVisible(False)
            self.NaliContainer.setVisible(False)
            self.PacketsContainer.setVisible(True)

        elif (self.itemListComboBox.currentText() == 'Jutta'):
            self.RollsContainer.setVisible(False)
            self.ReelsContainer.setVisible(False)
            self.PacketsContainer.setVisible(False)
            self.TotayContainer.setVisible(False)
            self.RaddiContainer.setVisible(False)
            self.NaliContainer.setVisible(False)
            self.JuttaContainer.setVisible(True)

        elif (self.itemListComboBox.currentText() == 'Totay'):
            self.RollsContainer.setVisible(False)
            self.ReelsContainer.setVisible(False)
            self.PacketsContainer.setVisible(False)
            self.JuttaContainer.setVisible(False)
            self.RaddiContainer.setVisible(False)
            self.NaliContainer.setVisible(False)
            self.TotayContainer.setVisible(True)
        elif (self.itemListComboBox.currentText() == 'Raddi'):
            self.RollsContainer.setVisible(False)
            self.ReelsContainer.setVisible(False)
            self.PacketsContainer.setVisible(False)
            self.JuttaContainer.setVisible(False)
            self.NaliContainer.setVisible(False)
            self.TotayContainer.setVisible(False)
            self.RaddiContainer.setVisible(True)

        elif (self.itemListComboBox.currentText() == 'Nali'):
            self.RollsContainer.setVisible(False)
            self.ReelsContainer.setVisible(False)
            self.PacketsContainer.setVisible(False)
            self.JuttaContainer.setVisible(False)
            self.RaddiContainer.setVisible(False)
            self.TotayContainer.setVisible(False)
            self.NaliContainer.setVisible(True)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Client Bill"))
        self.itemListComboBox.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.itemListComboBox.setItemText(1, _translate("MainWindow", "Rolls"))
        self.itemListComboBox.setItemText(2, _translate("MainWindow", "Reels"))
        self.itemListComboBox.setItemText(3, _translate("MainWindow", "Packets"))
        self.itemListComboBox.setItemText(4, _translate("MainWindow", "Jutta"))
        self.itemListComboBox.setItemText(5, _translate("MainWindow", "Totay"))
        self.itemListComboBox.setItemText(6, _translate("MainWindow", "Raddi"))
        self.itemListComboBox.setItemText(7, _translate("MainWindow", "Nali"))
        self.item_label.setText(_translate("MainWindow", "Item:"))
        self.item_label_35.setText(_translate("MainWindow", "Size:"))
        
        self.client_name_label.setText(_translate("MainWindow", "Client Name:"))
        self.RollsContainer.setTitle(_translate("MainWindow", "Rolls"))
        self.RollsContainer.setVisible(False)
        
        self.items.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.items.setItemText(1, _translate("MainWindow", "Tota"))
        self.items.setItemText(2, _translate("MainWindow", "Reel"))
        
        self.RollsPaperTypeComboBox.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.RollsPaperTypeComboBox.setItemText(1, _translate("MainWindow", "Fluting"))
        self.RollsPaperTypeComboBox.setItemText(2, _translate("MainWindow", "Fluting Bareek"))
        self.RollsPaperTypeComboBox.setItemText(3, _translate("MainWindow", "L1"))
        self.RollsPaperTypeComboBox.setItemText(4, _translate("MainWindow",  "L2"))
        self.RollsPaperTypeComboBox.setItemText(5, _translate("MainWindow", "L1 Bareek"))
        self.RollsPaperTypeComboBox.setItemText(6, _translate("MainWindow", "L2 Bareek"))
        self.RollsPaperTypeComboBox.setItemText(7, _translate("MainWindow", "Test Liner"))
        self.RollsPaperTypeComboBox.setItemText(8, _translate("MainWindow", "Test Liner Bareek"))
        self.RollsPaperTypeComboBox.setItemText(9, _translate("MainWindow", "Boxboard 2.5 No"))
        self.RollsPaperTypeComboBox.setItemText(10, _translate("MainWindow", "Boxboard 2.5 Bareek"))
        self.RollsPaperTypeComboBox.setItemText(11, _translate("MainWindow", "Boxboard 3 No"))
        self.RollsPaperTypeComboBox.setItemText(12, _translate("MainWindow", "Boxboard 3 Bareek"))
        self.RollsPaperTypeComboBox.setItemText(13, _translate("MainWindow", "Local Kraft"))
        self.RollsPaperTypeComboBox.setItemText(14, _translate("MainWindow", "Local Kraft Bareek"))
        self.RollsPaperTypeComboBox.setItemText(15, _translate("MainWindow", "Imported Kraft"))
        self.RollsPaperTypeComboBox.setItemText(16, _translate("MainWindow", "Imported Kraft Bareek"))
        self.RollsPaperTypeComboBox.setItemText(17, _translate("MainWindow", "Super Fluting"))
        self.RollsPaperTypeComboBox.setItemText(18, _translate("MainWindow", "Super Fluting Bareek"))
        
        self.item_label_3.setText(_translate("MainWindow", "Type:"))
        self.item_label_4.setText(_translate("MainWindow", "Quantity:"))
        self.item_label_5.setText(_translate("MainWindow", "Rate:"))
        self.Add_Rolls.setText(_translate("MainWindow", "Add"))
        self.item_label_19.setText(_translate("MainWindow", "Quantity in stock:"))
        self.item_label_7.setText(_translate("MainWindow", "Size:"))
        self.RollsSizeComboBox.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        
        self.TotaSizeTextField.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        for i in range(17,53):
                self.TotaSizeTextField.addItem(str(i))
                
        for i in range(17,53):
                self.RollsSizeComboBox.addItem(str(i))

        #self.item_label_8.setText(_translate("MainWindow", "Other Type"))
        
        self.packetsize.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        for i in range(17,53):
                self.packetsize.addItem(str(i))
        
        '''self.packetsize_totay.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        for i in range(17,53):
                self.packetsize_totay.addItem(str(i))'''
        self.ReelsContainer.setTitle(_translate("MainWindow", "Reels"))
        self.ReelsContainer.setVisible(False)
        self.item_label_11.setText(_translate("MainWindow", "Type:"))
        self.TotaWeightTextField.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.ReelsWeight.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.item_label_26.setText(_translate("MainWindow", "Rate:"))
        self.Add_Reels.setText(_translate("MainWindow", "Add"))
        self.item_label_27.setText(_translate("MainWindow", "Quantity in stock"))
        self.item_label_28.setText(_translate("MainWindow", "Size"))
        self.ReelsPaperTypeComboBox.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.ReelsPaperTypeComboBox.setItemText(1, _translate("MainWindow", "Fluting"))
        self.ReelsPaperTypeComboBox.setItemText(2, _translate("MainWindow", "L1"))
        self.ReelsPaperTypeComboBox.setItemText(3, _translate("MainWindow", "L2"))
        self.ReelsPaperTypeComboBox.setItemText(4, _translate("MainWindow", "TL"))
        self.ReelsPaperTypeComboBox.setItemText(5, _translate("MainWindow", "Kraft"))
        self.ReelsPaperTypeComboBox.setItemText(6, _translate("MainWindow", "Super Fluting"))
        self.ReelsPaperTypeComboBox.setItemText(7, _translate("MainWindow", "BB 2.5 No"))
        self.ReelsPaperTypeComboBox.setItemText(8, _translate("MainWindow", "BB Coated"))
        self.ReelsPaperTypeComboBox.setItemText(9, _translate("MainWindow", "BB 3 No"))
        self.item_label_30.setText(_translate("MainWindow", "Weight(kg):"))
        self.prev_balance.setText(_translate("MainWindow", "Current Balance:"))
        self.prevbalance.setText(_translate("MainWindow", (self.clientrecord[3] ))) 
        self.packetgram_m_label.setText(_translate("MainWindow","Grammage:") )
        self.JuttaContainer.setTitle(_translate("MainWindow", "Jutta"))
        self.JuttaContainer.setVisible(False)
        self.item_label_37.setText(_translate("MainWindow", "Weight (Kg):"))
        self.item_label_38.setText(_translate("MainWindow", "Rate:"))
        self.Add_Jutta.setText(_translate("MainWindow", u"Add", None))
        self.TotayContainer.setTitle(_translate("MainWindow", "Totay"))
        self.TotayContainer.setVisible(False)
        self.item_label_18.setText(_translate("MainWindow", "Size:"))
        self.item_label_22.setText(_translate("MainWindow", "Weight:"))
        self.item_label_34.setText(_translate("MainWindow", "Rate: "))
        self.Add_Totay.setText(_translate("MainWindow", "Add",None))
        self.PacketsContainer.setTitle(_translate("MainWindow", "Packets"))
        self.PacketsContainer.setVisible(False)
        self.item_label2.setText(_translate("MainWindow", "Length:"))
        self.Packetslength.setText('')
        self.item_label_20.setText(_translate("MainWindow", u"Quantity in Stock:", None))
                
        # self.item_label_15.setText(_translate("MainWindow", "Length="))
        self.item_label_16.setText(_translate("MainWindow", "Width:"))
        self.item_label_14.setText(_translate("MainWindow", "Number of Packets:"))
        self.item_label_13.setText(_translate("MainWindow", "Type:"))
        self.item_label_31.setText(_translate("MainWindow", "Rate:"))
        self.item_36.setText(_translate("MainWindow", "Tota/Reel:"))
        
        #self.item_label_35.setText(_translate("MainWindow", "Size:"))
        # self.packetsize.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.PacketsPaperTypeComboBox.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.PacketsPaperTypeComboBox.setItemText(1, _translate("MainWindow", "Fluting"))
        self.PacketsPaperTypeComboBox.setItemText(2, _translate("MainWindow", "L1"))
        self.PacketsPaperTypeComboBox.setItemText(3, _translate("MainWindow", "L2"))
        self.PacketsPaperTypeComboBox.setItemText(4, _translate("MainWindow", "TL"))
        self.PacketsPaperTypeComboBox.setItemText(5, _translate("MainWindow", "Kraft"))
        self.PacketsPaperTypeComboBox.setItemText(6, _translate("MainWindow", "Super Fluting"))
        self.PacketsPaperTypeComboBox.setItemText(7, _translate("MainWindow", "BB 2.5 No"))
        self.PacketsPaperTypeComboBox.setItemText(8, _translate("MainWindow", "BB Coated"))
        self.PacketsPaperTypeComboBox.setItemText(9, _translate("MainWindow", "BB 3 No")) 
        '''self.item_label_31.setText(_translate("MainWindow", "Rate:"))'''
        self.ReelsSizeComboBox.setItemText(0, QtCore.QCoreApplication.translate("MainWindow", u"select size from drop down", None))
        self.Add_Packets.setText(_translate("MainWindow", "Add"))
        self.item_label_8.setText(QtCore.QCoreApplication.translate("MainWindow", u"Debit Rs:", None))
        self.client_contact_label.setText(QtCore.QCoreApplication.translate("MainWindow", u"Client Contact No:", None))
        self.item_label_33.setText(_translate("MainWindow", "Weight:"))
        self.item_label_2.setText(_translate("MainWindow", "Reciept no:"))
        self.CheckOutButton.setText(_translate("MainWindow", "Proceed to Checkout"))
        self.rate_packets_12.setText(_translate("MainWindow", "Rent Rs:"))
        self.item_label_24.setText(_translate("MainWindow", "Credit Details:"))
        self.item_label_25.setText(_translate("MainWindow", "Credit Rs:"))
        self.item_label_6.setText(_translate("MainWindow", "Date:"))
        self.NaliContainer.setTitle(_translate("MainWindow", "Nali:"))
        self.NaliContainer.setVisible(False)
        self.NaliWeight.setText(_translate("MainWindow", "Weight:"))
        self.NaliRate.setText(_translate("MainWindow", "Rate:"))
        self.Add_Nali.setText(_translate("MainWindow", "Add"))
        self.item_label_23.setText(_translate("MainWindow", u"Add Weight Manually:", None))

        #self.item_label_20.setText(_translate("MainWindow", u"Quantity in Stock:", None))
        
        self.RaddiContainer.setTitle(_translate("MainWindow", "Raddi"))
        self.RaddiContainer.setVisible(False)
        self.RaddiWeight.setText(_translate("MainWindow", "Weight:"))
        self.RaddiRate.setText(_translate("MainWindow", "Rate:"))
        self.Add_Raddi.setText(_translate("MainWindow", "Add"))
        self.textBrowser.setHtml( QtCore.QCoreApplication.translate("MainWindow", u"<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:'MS Shell Dlg 2'; font-size:7.8pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:36pt; font-weight:600;\">  Ahmed Corrugation Machines</span></p></body></html>", None))
        self.CheckOutButton.setText( QtCore.QCoreApplication.translate("MainWindow", u"Proceed to Checkout", None))
        self.textBrowser_3.setHtml( QtCore.QCoreApplication.translate("MainWindow", u"<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:'MS Shell Dlg 2'; font-size:7.8pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>", None))
        self.groupBox.setTitle("")
        self.groupBox_2.setTitle("")
        self.RentTextField.setText('0')
        self.CreditTextField.setText('0')
        self.groupBox3.setTitle("")
        self.QuantityStockReels.setText('0')
        self.QuantityStockRolls.setText('0')
        self.RentTextField.setText(QtCore.QCoreApplication.translate("MainWindow", u"0", None))
        self.prev_rcp.setText(_translate("MainWindow", u"Previous Receipt:" , None)) 
        #self.prevrcp.setText(QtCore.QCoreApplication.translate("Form", u"0", None))    
        self.manualledger.setText(_translate("MainWindow", u"Ledger:", None))
        self.item_label_31.setText(_translate("MainWindow", u"Rate:", None))
        self.totaTypeComboBox_2.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.PacketsGrammage.setItemText(0, _translate("MainWindow", "Select from Drop Down"))
        self.totaTypeComboBox_2.setItemText(1, _translate("MainWindow", "Fluting"))
        self.totaTypeComboBox_2.setItemText(2, _translate("MainWindow", "L1"))
        self.totaTypeComboBox_2.setItemText(3, _translate("MainWindow", "L2"))
        self.totaTypeComboBox_2.setItemText(4, _translate("MainWindow", "TL"))
        self.totaTypeComboBox_2.setItemText(5, _translate("MainWindow", "Kraft"))
        self.totaTypeComboBox_2.setItemText(6, _translate("MainWindow", "Super Fluting"))
        self.totaTypeComboBox_2.setItemText(7, _translate("MainWindow", "BB 2.5 No"))
        self.totaTypeComboBox_2.setItemText(8, _translate("MainWindow", "BB Coated"))
        self.totaTypeComboBox_2.setItemText(9, _translate("MainWindow", "BB 3 No")) 
        def stock_out_func(dt,dets,typ,siz,qty,stockqty):
            global stock_out_rolls
            # qty is the quantity he just purchased for a size and item type whereas stockqty is the 
            # quantity left in stock after this buying operation
            new_row = {'date':dt, 'details': dets,'item_type':typ, 'size':siz, 'quantity' :qty     ,  'quantity_in_stock':stockqty    }
            #append row to the dataframe
            stock_out_rolls = stock_out_rolls.append(new_row, ignore_index=True) 

        def close_window():
            QtCore.QTimer.singleShot(0, MainWindow.close)
        
        def stock_out_func_tota(dt,dets,typ,siz,wght,rate):
             
            # Details = Stock Out Client Name
            new_row = {'date':dt, 'details': dets,'item_type':typ, 'size':siz, 'weight' :wght,  'rate':rate    }
            #append row to the dataframe
            self.stock_out_totay = self.stock_out_totay.append(new_row, ignore_index=True) 
 
        def write_excel(df, sheets, excel_path):
            book=openpyxl.load_workbook(excel_path )
            writer = pd.ExcelWriter(excel_path, engine='openpyxl',mode="a",if_sheet_exists="replace")
            writer.book = book
            writer.sheets = {ws.title:ws for ws in book.worksheets}
            result=pd.DataFrame()
            result = df
            result.to_excel(writer,sheet_name=sheets, index=False)
            writer.save() 
         
        def deleteentry():
            try:

                    current_row = self.cashbillDetailstable.currentRow()
                    current_column = self.cashbillDetailstable.currentColumn()
                    cell_value = self.cashbillDetailstable.item(current_row, current_column).text()  
                    
                    x = cell_value.split(' ')
                     
                    self.cashbillDetailstable.removeRow(self.cashbillDetailstable.currentRow())
                    self.updatedtext()
                    weight = ''
                    dates=''
                    '''global self.reelsstock
                    global self.rollsstock
                    global self.stock_out_totay'''
                    global vendorg
                    global details
                    if (x[-1].strip() == 'Rolls'):
                        quantity = int(x[-2])
                        papertype2 = str(x[1]).strip()
                        size = int(x[0])
                        det='Stock out '+self.clientName.toPlainText().strip()
                        papertype=papertype2.replace("_"," ")
                        
                        self.stock_out_rolls.drop(self.stock_out_rolls[(self.stock_out_rolls['details']==det )    &( self.stock_out_rolls['size']==int(float(size)))  & (self.stock_out_rolls['quantity']==int(float(quantity)) )& (self.stock_out_rolls['item_type']== (papertype))  ].index, inplace=True)   
                        if (papertype.lower().strip()=="Fluting".lower().strip() ):
                                '''global self.Fluting'''
                                for index, row in self.Fluting.iterrows():
                                    if   row[1] == size:
                                        row[2] += (quantity) 
                            
                        if (papertype.lower().strip()=="Fluting Bareek".lower().strip() ):

                            '''global self.Fluting_bareek'''
                            for index, row in self.Fluting_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="L1".lower().strip() ):

                            '''global self.L1'''

                            for index, row in self.L1.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="L1 Bareek".lower().strip() ):

                            '''global self.L1_bareek'''
                            for index, row in self.L1_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="L2".lower().strip() ):

                            '''global self.L2'''
                            for index, row in self.L2.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="L2 Bareek".lower().strip() ):

                            '''global self.L2_Bareek'''
                            for index, row in self.L2_Bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="Test Liner".lower().strip() ):

                            '''global self.testliner'''

                            for index, row in self.testliner.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Test Liner Bareek".lower().strip() ):

                            '''global self.testliner_bareek'''

                            for index, row in self.testliner_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Boxboard 2.5 No".lower().strip() ):

                            '''global self.boxboard2_5'''  
                            for index, row in self.boxboard2_5.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Boxboard2.5 Bareek".lower().strip() ):
                            '''global self.boxboard2_5_bareek'''

                            for index, row in self.boxboard2_5_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Boxboard 3 No".lower().strip() ):

                            '''global self.boxboard3 '''
                            for index, row in self.boxboard3.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="Boxboard 3 Bareek".lower().strip() ):

                            '''global self.boxboard3_bareek''' 
                            for index, row in self.boxboard3_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Local Kraft".lower().strip() ):

                            '''global self.localkraft '''
                            for index, row in self.localkraft.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="Local Kraft Bareek".lower().strip() ):

                            '''global self.localkraft_bareek'''

                            for index, row in self.localkraft_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="Imported Kraft".lower().strip() ):

                            '''global self.importedkraft'''

                            for index, row in self.importedkraft.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Imported Kraft Bareek".lower().strip() ):

                            '''global self.importedkraft_bareek'''
                            for index, row in self.importedkraft_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity)  

                        if (papertype.lower().strip()=="Super self.Fluting".lower().strip() ):

                            '''global self.Super_Fluting'''
                            for index, row in self.Super_Fluting.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 

                        if (papertype.lower().strip()=="Super self.Fluting Bareek".lower().strip() ):

                            '''global self.Super_Fluting_bareek'''
                            for index, row in self.Super_Fluting_bareek.iterrows():
                                if   row[1] == size:
                                        row[2] += (quantity) 
                            
                            
                            
                                
                    if (x[-1].strip() == 'Reel_Packets'):
                         
                         
                                    vend1="none"
                                    size1 = (x[0])
                                    size2 = size1.split('x')
                                    size = int(size2[0])
                                    papertype2 = str(x[1]).strip()
                                    weight1 = (x[-2])
                                    weight = int(float(weight1))
                                    
                                    papertype=papertype2.replace("_"," ")
                                    for i in range(len(vendorg)):
                                        if (vendorg[i][1]) ==papertype   and   (vendorg[i][2]) ==size  and   (vendorg[i][3]) == weight :
                                            vend1=(vendorg[i][4])#vendor
                                            dates=(vendorg[i][0])
                                        else:
                                            vend1="none"
                                            dates=''
                                            
                                    self.reelsstock.loc[len(self.reelsstock)] = [dates,papertype, size, weight,vend1,0]
                          #del row from the dataframe
                                    det='Stock out '+self.clientName.toPlainText().strip() 
                                    self.stock_out_reels.drop(self.stock_out_reels[(self.stock_out_reels['details']==det) & (self.stock_out_reels['date']==dates) &
                                                             ( self.stock_out_reels['size']==int(float(size)) ) &
                                                                                  (self.stock_out_reels['weight']==int(float(weight))) &
                                                                                  (self.stock_out_reels['item_type']== (papertype))].index, inplace=True)   

                               
       
                    if (x[-1].strip() == 'Tota_Packets'):
            
                         
                                    size1 = (x[0])
                                    size2 = size1.split('x')
                                    size = int(size2[0])
                                    papertype2 = str(x[1]).strip()
                                    weight1 = (x[-2])
                                    
                                    weight = int(float(weight1))
                                    details1="none"
                                    papertype=papertype2.replace("_"," ") 
                                    det='Stock out '+self.clientName.toPlainText().strip()
                                    for i in range(len(details)):
                                        if (details[i][1]) ==papertype   and   (details[i][2]) ==size  and   (details[i][3]) == weight :
                                            details1=(details[i][4])#vendor
                                            dates=(details[i][0])
                                        else:
                                            details1="none"
                                            dates= ''
                                    self.stock_out_totay.drop(self.stock_out_totay[(self.stock_out_totay['details']==det) &
                                                                         ( self.stock_out_totay['size']==int(float(size)) ) &
                                                                         (   self.stock_out_totay['date']==dates ) & 
                                                                                              (self.stock_out_totay['weight']==int(float(weight))) &
                                                                                              (self.stock_out_totay['item_type']== (papertype)) ].index, inplace=True)
                                    self.totaystock.loc[len(self.totaystock)] = [dates,papertype, size, weight,details1,0]


                                    

                    if (x[-1] == 'Reel'):
                        
                        #quantity = int(x[-2])
                        papertype2 = str(x[1]).strip()
                        size = int(x[0])
                        weight1 = x[2]
                        weight = int(float(weight1 ))
                        papertype=papertype2.replace("_"," ") 
                        for i in range(len(vendorg)):
                                        if (vendorg[i][1]) ==papertype   and   (vendorg[i][2]) ==size  and   (vendorg[i][3]) == weight :
                                            vend1=(vendorg[i][4])#vendor
                                            dates=(vendorg[i][0])
                                        else:
                                            vend1="none"
                                            dates=''
                        self.reelsstock.loc[len(self.reelsstock)] = [dates,papertype, size, weight,vend1,0]
                          #del row from the dataframe
                        
                        
                        det='Stock out '+self.clientName.toPlainText().strip()
                        self.stock_out_reels.drop(self.stock_out_reels[(self.stock_out_reels['details']==det) &
                                                             (   self.stock_out_reels['size']==int(float(size)) )  &
                                                             (   self.stock_out_reels['date']==dates ) & 
                                                                                  (self.stock_out_reels['weight']==int(float(weight))) &
                                                                                  (self.stock_out_reels['item_type']== (papertype))].index, inplace=True)
                       
               
                     
                    if (x[-1] == 'Tota'):
                        
                        #quantity = int(x[-2])
                        papertype2 = str(x[1]).strip()
                        size = int(x[0])
                        weight1 = x[2]
                        weight = int(float(weight1  ))
                        det='Stock out '+self.clientName.toPlainText().strip() 
                        papertype=papertype2.replace("_"," ")   
                        dates=''
                        for i in range(len(details)):
                                        if (details[i][1]) ==papertype   and   (details[i][2]) ==size  and   (details[i][3]) == weight :
                                            details1=(details[i][4])#vendor
                                            dates=(details[i][0])
                                        else:
                                            details1="none"
                                            dates= ''
                        self.stock_out_totay.drop(self.stock_out_totay[(self.stock_out_totay['details']==det) &
                                                             ( self.stock_out_totay['size']==int(float(size)) ) &
                                                             (   self.stock_out_totay['date']==dates ) & 
                                                                                  (self.stock_out_totay['weight']==int(float(weight))) &
                                                                                  (self.stock_out_totay['item_type']== (papertype)) ].index, inplace=True)
                        self.totaystock.loc[len(self.totaystock)] = [dates,papertype, size, weight,details1,0]
                     
                    self.reset_reels_container()
                    self.reset_jutta_container()
                    self.reset_totay_container()
                    self.reset_raddi_container()
                    self.reset_nali_container()
                    self.reset_packets_container()
                    self.reset_rolls_container()
                    
            except AttributeError:
                        msg = QMessageBox()  # create an instance of it
                        msg.setIcon(QMessageBox.Information)  # set icon
                        #msg.setWindowIcon(QtGui.QIcon("whatsapp-logo.png"))
                        msg.setText("No record selected")  # set text    
        
  
        
        def stock_out_func_reels(dt,dets,typ,siz,wght,rate):
            global stock_out_reels
            #Details = Stock Out Client Name
            new_row = {'date':dt, 'details': dets,'item_type':typ, 'size':siz, 'weight' :wght, 'rate':rate }
            #append row to the dataframe
            self.stock_out_reels = self.stock_out_reels.append(new_row, ignore_index=True) 

      
        def add_reels():
            # size  type qty  itemname
            '''global self.reelsstock'''
            errors = []
            if (self.ReelsSizeComboBox.currentText() == "Select from Drop Down"):
                errors.append("Size is not Selected")
            if (self.ReelsPaperTypeComboBox.currentText() == "Select from Drop Down"):
                errors.append('Papertype Not selected')
            if self.ReelsWeight.currentText() == "" or ((self.ReelsWeight.currentText() == "Select from Drop Down")):
                errors.append('Invalid Weight')
            if self.ReelsRate.toPlainText().strip() == "" or (not (bool(re.match('^\d+?\.\d+?|\d+$', self.ReelsRate.toPlainText().strip())))):
                errors.append('Wrong Input in Rate Text Field')
            if (len(errors) == 0):

                size = int(self.ReelsSizeComboBox.currentText().strip())
                itemtype1 = self.ReelsPaperTypeComboBox.currentText().strip()
                itemtype=itemtype1.replace(" ","_")
                #qty = int(self.ReelsQuantity.toPlainText().strip())
                itemname = 'Reel'
                weightitem = int(float(self.ReelsWeight.currentText()))
                rate = float(self.ReelsRate.toPlainText().strip())

                price = str(round(rate * weightitem,2))

                name = (str(size) + ' ' + str(itemtype) + ' ' + str(weightitem) + ' Kg '    + str(itemname))

                rowPosition = self.cashbillDetailstable.rowCount()
                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))

                self.updatedtext()

                weight = int(float(self.ReelsWeight.currentText()))
                
                '''global self.stock_out_reels'''
                    
                dets='Stock out '+self.clientName.toPlainText().strip()    
                #Details = Stock Out Client Name
                
                dt=str(date.today().strftime("%d-%m-%y") )
                
                stock_out_func_reels(dt,dets,itemtype1,size,weight,price) 
                global vendorg
                for index, row in self.reelsstock.iterrows():
                    if (row[1].strip().lower() == itemtype1.strip().lower()) &  ( row[2]==(int(self.ReelsSizeComboBox.currentText().strip()))) & ( (row[3])==int(weight)):
                        vendorg.append(row)
                        self.reelsstock.drop(index, inplace=True)
                        break
                self.reelsstock.reset_index(drop=True)
                self.reset_reels_container()

            else:

                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs = " , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text

                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message = msg.exec_()

        def add_rolls():
            '''global self.rollsstock'''
            errors = []
            if self.RollsSizeComboBox.currentText() == "Select from Drop Down":
                errors.append("Size is not Selected")
            if self.RollsPaperTypeComboBox.currentText() == "Select from Drop Down":
                errors.append("Papertype is not Selected")

            if (int(float(self.RollsQuantity.toPlainText().strip())) < 1) or (
            not (bool(re.match("^[0-9?]+$", self.RollsQuantity.toPlainText().strip())))):
                errors.append("invalid Quantity")
            if self.RollsRate.toPlainText().strip() == "" or (
            not (bool(re.match('^\d+?\.\d+?|\d+$', self.RollsRate.toPlainText().strip())))):
                errors.append("Wrong input in Rate Field") # set information under the main text
            if len(errors) == 0:
                size = int(self.RollsSizeComboBox.currentText())
                itemtype1 = str(self.RollsPaperTypeComboBox.currentText())
                itemtype=itemtype1.replace(" ","_")
                qty = int(self.RollsQuantity.toPlainText().strip())
                itemname = 'Rolls'
                rate = float(self.RollsRate.toPlainText().strip())
                price = str(round(rate * size * qty,2))
                name = (str(size) + ' ' + str(itemtype) + ' ' + str(qty) + ' ' + str(itemname))
                rowPosition = self.cashbillDetailstable.rowCount()
                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))
                self.updatedtext()
                '''global self.stock_out_rolls'''
                det='Stock out '+self.clientName.toPlainText().strip()    
                #Details = Stock Out Client Name
                
                dt=str(date.today().strftime("%d-%m-%y") )
                #self.rollsstock.reset_index(drop=True)
                if (itemtype1.lower().strip()=="Fluting".lower().strip() ):
                    '''global self.Fluting'''
                    for index, row in self.Fluting.iterrows():
                        if (  (row['Size'] == int(size))):
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                            newquantity = row['Quantity'] - (qty)
                            self.Fluting.loc[index, 'Quantity'] = newquantity 
                            
                if (itemtype1.lower().strip()=="Fluting Bareek".lower().strip() ):
                   
                    '''global self.Fluting_bareek'''
                    for index, row in self.Fluting_bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                            self.Fluting_bareek.loc[index, 'Quantity'] = newquantity 
                            
                if (itemtype1.lower().strip()=="L1".lower().strip() ):
                    
                    '''global self.L1'''
                     
                    for index, row in self.L1.iterrows():
                        if (  (row['Size'] == int(size))):
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                            newquantity = row['Quantity'] - (qty)
                            self.L1.loc[index, 'Quantity'] = newquantity 
                            
                if (itemtype1.lower().strip()=="L1 Bareek".lower().strip() ):
                    
                    '''global self.L1_bareek'''
                    for index, row in self.L1_bareek.iterrows():
                        if ((row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                            self.L1_bareek.loc[index, 'Quantity'] = newquantity 
                            
                if (itemtype1.lower().strip()=="L2".lower().strip() ):
                    
                    ''' global self.L2'''
                    for index, row in self.L2.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.L2.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="L2 Bareek".lower().strip() ):
                    
                    '''global self.L2_Bareek'''
                    for index, row in self.L2_Bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.L2_Bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Test Liner".lower().strip() ):
                    
                    '''global self.testliner'''
                     
                    for index, row in self.testliner.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.testliner.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Test Liner Bareek".lower().strip() ):
                    
                    '''global self.testliner_bareek'''
                    
                    for index, row in self.testliner_bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.testliner_bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Boxboard 2.5 No".lower().strip() ):
                    
                    '''global self.boxboard2_5'''  
                    for index, row in self.boxboard2_5.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.boxboard2_5.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Boxboard 2.5 Bareek".lower().strip() ):
                    
                    '''global self.boxboard2_5_bareek'''
                    
                    for index, row in self.boxboard2_5_bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.boxboard2_5_bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Boxboard 3 No".lower().strip() ):
                    
                    '''global self.boxboard3 '''
                    for index, row in self.boxboard3.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.boxboard3.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Boxboard 3 Bareek".lower().strip() ):
                    
                    '''global self.boxboard3_bareek''' 
                    for index, row in self.boxboard3_bareek.iterrows():
                        if ((row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.boxboard3_bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Local Kraft".lower().strip() ):
                    
                    '''global self.localkraft''' 
                    for index, row in self.localkraft.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.localkraft.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Local Kraft Bareek".lower().strip() ):
                    
                    '''global self.localkraft_bareek'''
                   
                    for index, row in self.localkraft_bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.localkraft_bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Imported Kraft".lower().strip() ):
                    
                    '''global self.importedkraft'''
                     
                    for index, row in self.importedkraft.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.importedkraft.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Imported Kraft Bareek".lower().strip() ):
                    
                    '''global self.importedkraft_bareek'''
                    for index, row in self.importedkraft_bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.importedkraft_bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Super self.Fluting".lower().strip() ):
                    
                    '''global self.Super_Fluting'''
                    for index, row in self.Super_Fluting.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.Super_Fluting.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                if (itemtype1.lower().strip()=="Super self.Fluting Bareek".lower().strip() ):
                    
                    '''global self.Super_Fluting_bareek'''
                    for index, row in self.Super_Fluting_bareek.iterrows():
                        if (  (row['Size'] == int(size))):
                            newquantity = row['Quantity'] - (qty)
                            self.Super_Fluting_bareek.loc[index, 'Quantity'] = newquantity 
                            stock_out_func(dt,det,itemtype1,size,qty,row['Quantity'])
                     
                self.reset_rolls_container()
            else:

                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs = " , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text

                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message = msg.exec_()

        def add_packets():
            '''global self.reelsstock
            global self.totaystock'''
            errors = []
            
            if ( not (bool(re.match('\d+$', self.packetgram_m.toPlainText().strip()))))  :
                errors.append('Wrong input in Grammage Text Field')
            
            if self.Packetslength.toPlainText().strip() == "" or (not (bool(re.match('\d+$', self.Packetslength.toPlainText().strip())))):
                errors.append('Wrong Input in Packets Length Field ')
             
            if self.PacketsWidth.toPlainText().strip() == "" or (not (bool(re.match('\d+$', self.PacketsWidth.toPlainText().strip())))):
                errors.append('Wrong Input in Packets Width Field ')
            if self.PacketsPaperTypeComboBox.currentText() == "Select from Drop Down":
                errors.append('Papertype Not Selected')

            if self.PacketsNoOfPackets.toPlainText().strip() == "" or (not (bool(re.match('^\d+$', self.PacketsNoOfPackets.toPlainText().strip())))):
                errors.append('Wrong Input in No. of Packets Field')

            if (self.PacketsRate.toPlainText().strip() == "") or (not (bool(re.match('^\d+?\.\d+?|\d+$', self.PacketsRate.toPlainText().strip())))):
                errors.append("Wrong Input in Rate Text Field")

            
            if len(errors) == 0:
                grams=self.packetgram_m.toPlainText().strip()
                               
                length2=           self.Packetslength.toPlainText().strip()    
                choice=     self.items.currentText().strip()  
                               
                if choice=='Reel':
                               itemname='Reel_Packets'
                elif choice=='Tota':
                               itemname="Tota_Packets"
                else:
                    itemname="Packets"
                               
                if (self.packetsize.currentText() != "Select from Drop Down"): 
                    length=self.packetsize.currentText().strip()
                     
                 
                size = int(length2) * int(self.PacketsWidth.toPlainText().strip())* int((grams))
                itemtype1 = self.PacketsPaperTypeComboBox.currentText().strip()
                itemtype=itemtype1.replace(" ","_")
                qty = int(self.PacketsNoOfPackets.toPlainText().strip())
                rate = float(self.PacketsRate.toPlainText().strip())
                price = str(round((size / 15500) * qty * rate,2))
                grammage = int(grams) 
                '''if (self.items.currentText() != "Select from Drop Down"): 
                    name=(str(length2) + 'x' + str(self.PacketsWidth.toPlainText().strip()) + ' ' + str(itemtype) + ' ' + str(grammage) + ' kg ' + str(qty) + ' ' + str(itemname))
                else:'''
                name=(str(length2) + 'x' + str(self.PacketsWidth.toPlainText().strip()) + ' ' + str(itemtype) + ' '  + str(qty) + ' ' + str(itemname))
                
                rowPosition = self.cashbillDetailstable.rowCount()
                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))
                self.updatedtext()
                grammage = int(grams) 
                global vendorg
                global details 
                if (self.items.currentText() != "Select from Drop Down"):
                    if (self.items.currentText().strip().lower() == "reel".strip()): 
                        '''global self.stock_out_reels'''
                        dets='Stock out '+self.clientName.toPlainText().strip()  
                        #dates= (self.dateTimeEdit.date().toPyDate().strftime("%Y-%m-%d"))
                        
                        dt=str(date.today().strftime("%d-%m-%y") )
                        
                        stock_out_func_reels(dt,dets,itemtype1,size,grammage,price) 
                        if (self.packetsize.currentText() != "Select from Drop Down"):
                            length=self.packetsize.currentText().strip()
                            for index, row in self.reelsstock.iterrows():
                                if row[1].strip().lower() == itemtype1.strip().lower() and  row[2] == int( (length))  and  row[3] ==  (grammage) :
                                    vendorg.append(row)
                                    self.reelsstock.drop(index, inplace=True)
                            self.reelsstock.reset_index(drop=True)
                    if (self.items.currentText().strip().lower() == "tota"):
                        '''global self.stock_out_totay''' 
                        dets='Stock out '+self.clientName.toPlainText().strip() 
                                    #Details = Stock Out Client Name
                        
                        dt=str(date.today().strftime("%d-%m-%y") )
                         
                        stock_out_func_tota(dt,dets,itemtype1,size,grammage,price) 
                        
                        if (self.packetsize.currentText() != "Select from Drop Down"): 
                                length=self.packetsize.currentText().strip()
                                for index, row in self.totaystock.iterrows():
                                    if row[1].strip().lower() == itemtype1.strip().lower()  and  row[2] == int( (length))  and  row[3]  ==  ( (grammage)) :
                                        details.append(row)
                                        self.totaystock.drop(index, inplace=True)
                                self.totaystock.reset_index(drop=True)
    
                
                    self.reset_packets_container()
                    self.items.clear()
                    self.items.addItem("Select from Drop Down")
                    self.items.addItem("Reel")
                    self.items.addItem("Tota")
                else:

                    msg = QMessageBox()  # create an instance of it
                    msg.setIcon(QMessageBox.Information)  # set icon
                    msgs = " , ".join([str(item) for item in errors])
                    msg.setText(msgs)  # set text

                    msg.setWindowTitle("Alert")  # set title
                    message = msg.exec_()


        def add_jutta():
            errors = []
            if (self.JuttaWeightTextField.toPlainText().strip() == ''):
                errors.append("weight text field is empty")

            if self.JuttaRateTextField.toPlainText().strip() == '':
                errors.append('rate text field is empty')

            if not (bool(re.match('^\d+?\.\d+?|\d+$', self.JuttaWeightTextField.toPlainText().strip()))):
                errors.append("wrong input in Jutta weight text field")

            if not (bool(re.match('^\d+?\.\d+?|\d+$', self.JuttaRateTextField.toPlainText().strip()))):
                errors.append("wrong input in Jutta rate text field")

            if (len(errors) == 0):

                weight = int(float(self.JuttaWeightTextField.toPlainText().strip()))

                itemname = 'Jutta'
                rate = int(self.JuttaRateTextField.toPlainText().strip())
                price = str(round(rate * weight,2))
                name = (str(weight) + ' Kg ' + str(itemname))
                rowPosition = self.cashbillDetailstable.rowCount()
                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))
                self.updatedtext()
                self.reset_jutta_container()
            else:

                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs = " , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text

                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message = msg.exec_()

        def add_totay():
            errors = []
            if self.TotaRateTextField.toPlainText().strip() == '': errors.append(" rate text field is empty")
            if not (bool(re.match('^[0-9]+$', self.TotaRateTextField.toPlainText().strip()))):
                errors.append("wrong input in rate text field")
            '''global self.totaystock'''
            if (self.totaTypeComboBox_2.currentText() == "Select from Drop Down"):
                errors.append('Type Not Selected')
            if (self.TotaWeightTextField.currentText() != "Select from Drop Down") and (self.TotaWeightTextField_2.toPlainText().strip()!=""):
                errors.append('Either enter grams manually or using drop down. You selected both options')
            if (self.TotaWeightTextField.currentText() == "Select from Drop Down") and   (not (bool(re.match('\d+$', self.TotaWeightTextField_2.toPlainText().strip()))))  :
                errors.append('Wrong input in Weight Text Field')
            
            if self.totaquantity.toPlainText().strip() == "" or (not (bool(re.match('^\d+$', self.totaquantity.toPlainText().strip())))):
                errors.append('Wrong input in Quantity field')
                
            if self.TotaSizeTextField.currentText() == "Select from Drop Down":
                errors.append('Tota Size/length is not selected')
            if (self.TotaRateTextField.toPlainText().strip() == "") or (not (bool(re.match('^\d+?\.\d+?|\d+$', self.TotaRateTextField.toPlainText().strip())))):
                errors.append("Wrong input in Rate Text Field")
            if (len(errors) != 0):
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs = " , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text
                msg.setWindowTitle("Alert")  # set title
                message = msg.exec_()
            else:
                if (self.TotaWeightTextField.currentText() != "Select from Drop Down"): 
                    weight=self.TotaWeightTextField.currentText().strip()
                else:
                    weight=self.TotaWeightTextField_2.toPlainText().strip()
                itemname = 'Tota'
                itemtype1=self.totaTypeComboBox_2.currentText().strip()
                itemtype=itemtype1.replace(" ","_")
                size = str(self.TotaSizeTextField.currentText().strip())
                rate = float(self.TotaRateTextField.toPlainText().strip())
                price = str(round(rate * int(float(weight)),2))
                name = (str(size) + ' ' + str(itemtype) + ' ' +str(weight) + ' Kg ' + str(itemname))
                rowPosition = self.cashbillDetailstable.rowCount()
                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))
                self.updatedtext()
                
                #
                '''global self.stock_out_totay''' 
                global details 
                
                dets='Stock out '+self.clientName.toPlainText().strip()+' Cash Customer'    
                                    #Details = Stock Out Client Name
                #dt= str(date.today()  .toPyDate().strftime("%Y-%m-%d"))
                #e = datetime.now()
                dt=str(date.today().strftime("%d-%m-%y") )
                stock_out_func_tota(dt,dets,itemtype1,size,weight,price) 
                for index, row in self.totaystock.iterrows():
                    if row[0].strip().lower()==self.totaTypeComboBox_2.currentText().strip().lower() and row[1]==int(size) and row[2]==int(weight):
                            details.append(row)
                            self.totaystock.drop(index, inplace=True)
                self.totaystock.reset_index(drop=True)
                self.reset_totay_container()

        def add_raddi():
            errors = []

            if self.RaddiWeightTextField.toPlainText().strip() == '': errors.append("Raddi Weight text field is empty")
            if self.RaddiRateTextField.toPlainText().strip() == '': errors.append("Raddi Rate text field is empty")
            if not (bool(re.match('\d+$', self.RaddiWeightTextField.toPlainText().strip()))):
                errors.append("wrong input in weight text field")

            if not (bool(re.match('^\d+$', self.RaddiRateTextField.toPlainText().strip()))):
                errors.append("wrong input in raddi rate text field")
            if len(errors) == 0:
                weight = int(self.RaddiWeightTextField.toPlainText().strip())
                itemname = 'Raddi'
                rate = float(self.RaddiRateTextField.toPlainText().strip())
                price = str(round(rate * weight,2))
                name = (str(itemname) + ' ' + str(weight) + ' Kg')
                rowPosition = self.cashbillDetailstable.rowCount()
                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))
                self.updatedtext()
                self.reset_raddi_container()
            if (len(errors) != 0):
                #print(errors, " erors ")
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs = " , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text

                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message = msg.exec_()

        def Add_Nali():
            errors = []

            if self.NaliWeightTextField.toPlainText().strip() == '':
                errors.append("weight text field is empty")

            if not (bool(re.match('\d+$', self.NaliWeightTextField.toPlainText().strip()))):
                errors.append("wrong input in weight text field")

            if not (bool(re.match('^\d+?\.\d+?|\d+$', self.NaliRateTextField.toPlainText().strip()))):
                errors.append("wrong input in nali rate text field")

            if self.NaliRateTextField.toPlainText().strip() == '':
                errors.append("nali rate text field is empty")

            if len(errors) == 0:
                weight = int(self.NaliWeightTextField.toPlainText().strip())
                itemname = 'Nali'
                rate = float(self.NaliRateTextField.toPlainText().strip())
                price = str(round(rate * weight,2))

                name = (str(itemname) + ' ' + str(weight) + ' Kg ')
                rowPosition = self.cashbillDetailstable.rowCount()

                self.cashbillDetailstable.insertRow(rowPosition)
                self.cashbillDetailstable.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(name))
                self.cashbillDetailstable.setItem(rowPosition, 1, QtWidgets.QTableWidgetItem(price))
                self.updatedtext()
                self.reset_nali_container()
            if (len(errors) != 0):
                #print(errors, " erors ")
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs = " , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text

                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message = msg.exec_()

        def quantityrollsstockcheck():
            '''global self.rollsstock'''
            if (self.RollsPaperTypeComboBox.currentText() == "Select from Drop Down" or self.RollsSizeComboBox.currentText() == "Select from Drop Down"):
                self.QuantityStockRolls.setText('0')

            else:
                itemtype = (self.RollsPaperTypeComboBox.currentText()).strip()
                size = int(self.RollsSizeComboBox.currentText())
                qty = ''
                if (itemtype.lower().strip()=="Fluting".lower().strip() ):                    
                    '''global self.Fluting'''
                    #self.Fluting['Item_Type'] = self.rollsstock['Item_Type'].astype(str)
                    res = self.Fluting[ (self.Fluting['Size'] == size)]
                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty
                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)
                    else:
                        self.QuantityStockRolls.setText('0')
                if (itemtype.lower().strip()=="Fluting Bareek".lower().strip() ):
                   
                    '''global self.Fluting_bareek'''
                    res = self.Fluting_bareek[ (self.Fluting_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="L1".lower().strip()):
                    
                    '''global self.L1'''
                    res = self.L1[ (self.L1['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                     
                            
                if (itemtype.lower().strip()=="L1 Bareek".lower().strip() ):
                    
                    '''global self.L1_bareek'''
                    res = self.L1_bareek[ (self.L1_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="L2".lower().strip() ):
                    
                    '''global self.L2'''
                    res = self.L2[ (self.L2['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="L2 Bareek".lower().strip() ):
                    
                    '''global self.L2_Bareek'''
                    res = self.L2_Bareek[ (self.L2_Bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="Test Liner".lower().strip() ):
                    
                    '''global self.testliner'''
                    res = self.testliner[ (self.testliner['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                     
                            
                if (itemtype.lower().strip()=="Test Liner Bareek".lower().strip() ):
                    
                    '''global self.testliner_bareek'''
                    res = self.testliner_bareek[ (self.testliner_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0')
                     
                            
                if (itemtype.lower().strip()=="Boxboard 2.5 No".lower().strip() ):
                    '''global self.boxboard2_5''' 
                    
                    res = self.boxboard2_5[ (self.boxboard2_5['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0')
                    
                     
                            
                if (itemtype.lower().strip()=="Boxboard 2.5 Bareek".lower().strip() ):
                    
                    '''global self.boxboard2_5_bareek'''
                    res = self.boxboard2_5_bareek[ (self.boxboard2_5_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0')
                     
                            
                if (itemtype.lower().strip()=="Boxboard 3 No".lower().strip() ):
                    
                    '''global self.boxboard3 '''
                    res = self.boxboard3[(self.boxboard3['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="Boxboard 3 Bareek".lower().strip() ):
                    
                    ''' global self.boxboard3_bareek '''
                    res = self.boxboard3_bareek[ (self.boxboard3_bareek['Size'] == size)]
                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty
                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)
                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="Local Kraft".lower().strip() ):
                    
                    '''global self.localkraft''' 
                    res = self.localkraft[ (self.localkraft['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="Local Kraft Bareek".lower().strip() ):
                    
                    '''global self.localkraft_bareek'''
                   
                    res = self.localkraft_bareek[ (self.localkraft_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="Imported Kraft".lower().strip() ):
                    
                    '''global self.importedkraft'''
                    res = self.importedkraft[ (self.importedkraft['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                     
                            
                if (itemtype.lower().strip()=="Imported Kraft Bareek".lower().strip() ):
                    
                    '''global self.importedkraft_bareek'''
                    res = self.importedkraft_bareek[ (self.importedkraft_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                            
                if (itemtype.lower().strip()=="Super self.Fluting".lower().strip() ):
                    
                    '''global self.Super_Fluting'''
                    res = self.Super_Fluting[ (self.Super_Fluting['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0')
                if (itemtype.lower().strip()=="Super self.Fluting Bareek".lower().strip() ):
                    
                    '''global self.Super_Fluting_bareek'''
                    res = self.Super_Fluting_bareek[ (self.Super_Fluting_bareek['Size'] == size)]


                    qty = ((res['Quantity'].to_string(index=False)))
                    isempty = res.empty

                    if (isempty==False):
                        self.QuantityStockRolls.setText(qty)

                    else:
                        self.QuantityStockRolls.setText('0') 
                    
                    
                     
        def quantityreelsweightstockcheck():
            '''global self.reelsstock'''
            if (self.ReelsPaperTypeComboBox.currentText() == "Select from Drop Down" or self.ReelsSizeComboBox.currentText() == "Select from Drop Down"):
                self.QuantityStockReels.setText('0')
                self.ReelsWeight.clear()
                self.ReelsWeight.addItem("Select from Drop Down")
            if (self.ReelsPaperTypeComboBox.currentText() != "Select from Drop Down") and (self.ReelsSizeComboBox.currentText() != "Select from Drop Down"):
                if (self.ReelsSizeComboBox.currentText() != ""):

                    papertype = (self.ReelsPaperTypeComboBox.currentText())
                    size = int((self.ReelsSizeComboBox.currentText()))
                    qty = ''
                    self.reelsstock['Item_Type'] = self.reelsstock['Item_Type'].astype(str)
                    res = self.reelsstock[(self.reelsstock['Item_Type'].str.strip() == papertype.strip()) & (self.reelsstock['Size'] == size)]
                    qty = str(len(res))
                    self.QuantityStockReels.setText(qty)
                    weights = (res['Weight_g'].unique())
                    if (len(weights) != 0):
                        self.ReelsWeight.clear()
                        self.ReelsWeight.addItem("Select from Drop Down")
                        for i in weights:
                            self.ReelsWeight.addItem(str(i))

        def quantityreelssizestockcheck():

            '''global self.reelsstock'''

            if not (self.ReelsPaperTypeComboBox.currentText() == "Select from Drop Down"):
                selected = self.ReelsPaperTypeComboBox.currentText()
                res = self.reelsstock[(self.reelsstock['Item_Type'].str.strip() == selected.strip())]
                sizes= res['Size'].unique()  
                if (len(sizes) != 0):
                    self.ReelsSizeComboBox.clear()
                    self.ReelsSizeComboBox.addItem("Select from Drop Down")
                    sizes.sort()
                    for i in sizes:
                        self.ReelsSizeComboBox.addItem(str(i))
            else:
                self.ReelsWeight.clear()
                self.ReelsSizeComboBox.clear()
                self.ReelsSizeComboBox.addItem("Select from Drop Down")
                self.ReelsWeight.addItem("Select from Drop Down")
                self.QuantityStockReels.setText('0')

        def packetssizelist():
            '''global self.reelsstock'''
             
            if not (self.PacketsPaperTypeComboBox.currentText() == "Select from Drop Down"):
                selected = self.PacketsPaperTypeComboBox.currentText()
                res = self.reelsstock[(self.reelsstock['Item_Type'].str.strip() == selected.strip())]
                res.sort_values(by=['Size'])
                sizes= res['Size'].unique()  
                if (len(sizes) != 0):
                    self.packetsize.clear()
                    self.packetsize.addItem("Select from Drop Down")
                    sizes.sort()
                    for i in sizes:
                        self.packetsize.addItem(str(i))
            else:
                self.PacketsGrammage.clear()
                self.packetsize.clear()
                self.packetsize.addItem("Select from Drop Down")
                self.PacketsTotaDetails.setText('0')
                
        def totaypacketquantityweightlist():
            '''global self.totaystock
            global self.reelsstock'''
            if (self.PacketsPaperTypeComboBox.currentText() == "Select from Drop Down" or self.packetsize.currentText() == "Select from Drop Down" or self.items.currentText().strip().lower() == "select from drop down"    ):
                self.PacketsTotaDetails.setText('packettota')
                #self.packetgram_m.setText("")
                '''self.PacketsGrammage.clear()
                self.PacketsGrammage.addItem("Select from Drop Down")'''
            
        
                           #or self.packetsize_totay.currentText() == "Select from Drop Down"
                
            if (self.PacketsPaperTypeComboBox.currentText() != "Select from Drop Down") and ( self.packetsize.currentText() != "Select from Drop Down")  and ( self.items.currentText() != "Select from Drop Down"):
                if (self.items.currentText().strip().lower()== "tota"):
                    papertype = (self.PacketsPaperTypeComboBox.currentText().strip() )
                    size = float(self.packetsize.currentText())
                    qty = ''
                    self.totaystock['Item_Type'] = self.totaystock['Item_Type'].astype(str)
                    res = self.totaystock[(self.totaystock['Item_Type'].str.strip()  == papertype.strip())  &  (self.totaystock['Size'] == size)]
                    qty = str(len(res))
                    import traceback
                    try:
                        self.PacketsTotaDetails.setText(qty)
                    except RuntimeError:
                                     traceback.print_exc()   
                
                    weights = (res['Weight_g'].unique())
                    if (len(weights) != 0):
                        self.PacketsGrammage.clear()
                        
                        self.PacketsGrammage.addItem("Select from Drop Down")
                        for i in weights:
                            self.PacketsGrammage.addItem(str(i)) 
                               
                if (self.items.currentText().strip().lower()== "reel"):
                        papertype = (self.PacketsPaperTypeComboBox.currentText())
                        size = float(self.packetsize.currentText())
                        qty = ''
                        self.reelsstock['Item_Type'] = self.reelsstock['Item_Type'].astype(str)
                        res = self.reelsstock[
                            (self.reelsstock['Item_Type'].str.strip() == papertype.strip())  &  (self.reelsstock['Size'] == size)]
                        qty = str(len(res))
                        self.PacketsTotaDetails.setText(qty)
                        weights = (res['Weight_g'].unique())
                        if (len(weights) != 0):
                            self.PacketsGrammage.clear()
                            self.PacketsGrammage.addItem("Select from Drop Down")
                            for i in weights:
                                self.PacketsGrammage.addItem(str(i))                   
            else:  
                        self.PacketsGrammage.clear()
                        
                        self.PacketsGrammage.addItem("Select from Drop Down")
 
        def ratelistrolls():
            if (not(self.RollsPaperTypeComboBox.currentText() == "Select from Drop Down" )):
                itemtype = (self.RollsPaperTypeComboBox.currentText()).strip().lower()
                '''global self.rollsstock'''
                
                for index,row in self.rollsstock.iterrows():
                        if (row['Type'].strip().lower()==itemtype):
                            self.RollsRate.setText(str(row['Rate']))
            else: 
                self.RollsRate.setText("")
                
                 
        
        def totay_quantityweightlist():
            '''global self.totaystock'''
            if (self.totaTypeComboBox_2.currentText() == "Select from Drop Down" or self.TotaSizeTextField.currentText() == "Select from Drop Down" ):
                self.totayquantityfield.setText('0')
                self.totaquantity.setText('0')
                self.TotaWeightTextField_2.setText('0')
                self.TotaWeightTextField.clear()
                self.TotaWeightTextField.addItem("Select from Drop Down") 
            if (self.totaTypeComboBox_2.currentText() != "Select from Drop Down") and ( self.TotaSizeTextField.currentText() != "Select from Drop Down"):
                    papertype = (self.totaTypeComboBox_2.currentText())
                    size = float(self.TotaSizeTextField.currentText())
                    qty = ''
                    self.totaystock['Item_Type'] = self.totaystock['Item_Type'].astype(str)
                    res =self.totaystock[(self.totaystock['Item_Type'].str.strip() == papertype.strip())  &  (self.totaystock['Size'] == size)]
                    qty = str(len(res))
                    self.totayquantityfield.setText(qty)
                    weights = (res['Weight_g'].unique())
                    if (len(weights) != 0):
                        self.TotaWeightTextField.clear()
                        self.TotaWeightTextField.addItem("Select from Drop Down")
                        for i in weights:
                            self.TotaWeightTextField.addItem(str(i))
                    else:  
                        self.TotaWeightTextField.clear()
                        self.TotaWeightTextField.addItem("Select from Drop Down")
             
        self.Add_Reels.clicked.connect(add_reels)
        self.Add_Rolls.clicked.connect(add_rolls)
        self.Add_Packets.clicked.connect(add_packets)
        self.Add_Jutta.clicked.connect(add_jutta)
        self.Add_Totay.clicked.connect(add_totay)
        self.Add_Raddi.clicked.connect(add_raddi)
        self.Add_Nali.clicked.connect(Add_Nali)
        self.DELETEROWS.clicked.connect(deleteentry)
        self.RollsSizeComboBox.currentTextChanged.connect(quantityrollsstockcheck)
        self.RollsPaperTypeComboBox.currentTextChanged.connect(ratelistrolls)
        self.ReelsSizeComboBox.currentTextChanged.connect(quantityreelsweightstockcheck)
        self.ReelsPaperTypeComboBox.currentTextChanged.connect(quantityreelssizestockcheck)
        #self.PacketsPaperTypeComboBox.currentTextChanged.connect(packetssizelist)
        self.packetsize.activated.connect(totaypacketquantityweightlist)  # totay packet both
        self.TotaSizeTextField.currentTextChanged.connect(totay_quantityweightlist)
        
        def whtsapp(customernumber,reciept_number,customername):
            pth = 'Cash_Bills\\'
            pdf_name = str(customername) + "_" + str(reciept_number) + "_" +str(customernumber)+".pdf"          
            path = pth+pdf_name
            path2 = pth
            images = convert_from_path(path,poppler_path=r'poppler-0.68.0\bin')
            customer_number = '+'+str(customernumber)
            customer_name = str(customername)
            message = "Hello Mr. " + str(customer_name) + " here is your bill for the date: " + str(date.today()) + ". Thank you for purchasing from us, please come back again or call us at +923244023811"
            for i in range(len(images)):
                # Save pages as images in the pdf
                    images[i].save(path2+'\\''page'+ str(i) +'.jpg', 'JPEG')
                    pywhatkit.sendwhats_image(customer_number, str('page'+ str(i) +'.jpg'),message,wait_time=30)

        
        def cashbill():
            errors=[] 
            clients=self.clients 
            CLIENT_NAME =self.clientName.toPlainText().strip() 
            CONTACT_NO=self.contact.toPlainText().strip()
            CLIENT_ID= ( self.clientrecord[0])     
            ''' self.prevbalance.setText(self.clientrecord[3])
            self.prevrcp.setText(self.clients[1]) ''' 
            ledger=self.m_ledger.toPlainText().strip()
            if  self.CreditDetailsTextField.toPlainText().strip() == "":
                errors.append("Wrong input in Credit Details Text Field")
            if (ledger == "") or (not (bool(re.match('^\d+$', ledger)))):
                errors.append("Wrong Input in ledger Text Field")    
            if (self.CreditTextField.toPlainText().strip() == "" )or (not (bool(re.match('^\d+?\.\d+?|\d+$',self.CreditTextField.toPlainText().strip())))):
                errors.append("Wrong Input in Credit Text Field")
            if (not (bool(re.match('^\d+?\.\d+?|\d+$', self.RentTextField.toPlainText().strip())))) or (self.RentTextField.toPlainText().strip() == ''):
                errors.append("Wrong Input in Rent Text Field")            
            '''if (self.contact.toPlainText().strip() == '') or (not bool(re.match('^[1-9]\d{11,12}$' , self.contact.toPlainText().strip()))):
                errors.append('Wrong Input in Client Contact Text Field')'''
            if (not (bool(re.match("^[a-zA-z]+([\s][a-zA-Z]+)*$", self.clientName.toPlainText().strip())))) or (
                    self.clientName.toPlainText().strip() == ''):
                errors.append('Wrong Input in Client Name Text Field ')
            if (len(errors) != 0  ):
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msgs=" , ".join([str(item) for item in errors])
                msg.setText(msgs)  # set text
                
                '''msg.setInformativeText()'''  # set information under the main text
                msg.setWindowTitle("Alert")  # set title
                message=msg.exec_()
            if(len(errors)==0):
                totalrows = self.cashbillDetailstable.rowCount()
                updated1 = ''

                if totalrows != 0:
                    for i in range(0, totalrows):
                        if (i == totalrows - 1):
                            updated1 = updated1 + str(self.cashbillDetailstable.item(i, 0).text())
                        else:
                            updated1 = updated1 + str(self.cashbillDetailstable.item(i, 0).text()) + ' || '

                
                DATE= str(date.today().strftime("%d-%m-%y") )
                #existCLIENT =  clients [clients['CONTACT_NO'] == CONTACT_NO]
                 
                
                updated=updated1.replace(",", " ");
                 
                
                DETAILS_OF_BILL=str(updated)

                RECIEPT_NUMBER=self.rcp_no.toPlainText().strip()
                DEBIT=float(self.totalTextField.toPlainText().strip())+ float(self.RentTextField.toPlainText().strip())
                RENT=self.RentTextField.toPlainText().strip()
                CREDIT= (self.CreditTextField.toPlainText().strip())
                CREDIT_DETAILS=self.CreditDetailsTextField.toPlainText().strip()
                 
                text =CLIENT_ID
                 
                if (len(clients)<=1 and clients['RECIEPT_NUMBER'][0]==0):
                
                    BALANCE=str(round(float(DEBIT)-float(CREDIT),2))
                     
                    clients.loc[0] =[DATE, RECIEPT_NUMBER, CLIENT_ID, CLIENT_NAME, CONTACT_NO, DETAILS_OF_BILL, DEBIT, CREDIT, CREDIT_DETAILS, RENT, BALANCE,ledger]
                else:
                    previousbalance=clients['BALANCE'].loc[len(clients)-1]
                    BALANCE=round(float(previousbalance)+(float(DEBIT)-float(CREDIT)),2)
                    
                    clients.loc[len(clients)] =[DATE, RECIEPT_NUMBER, CLIENT_ID, CLIENT_NAME, CONTACT_NO, DETAILS_OF_BILL, DEBIT, CREDIT, CREDIT_DETAILS, RENT, BALANCE,ledger]

                    
                write_excel(clients,text,r'book.xlsx' )
                  
                 
                 
                 
                global pdf_date
                global pdf_client_name
                global pdf_client_contact_number
                global pdf_reciept_number
                global pdf_debit
                global pdf_credit
                global pdf_credit_details
                global pdf_rent
                global pdf_previous_balance
                global pdf_previous_bill_number
                global pdf_total_balance
                global pdf_dataframe
                clientssheets=pd.read_excel('book.xlsx', index_col=None,sheet_name=None)
                toremove=['reels_stock_in_out','tota_stock_in_out','rolls_stock_in_out','client info','rolls_stock','reels_stock','totay','Fluting', 'Fluting_Bareek', 'L1', 'L1_Bareek', 'L2', 'L2_Bareek', 'Test_Liner', 'Test_Liner_Bareek', 'Box_Board_2_5_No', 'Box_Board_2_5_No_Bareek', 'Box_Board_3_No', 'Box_Board_3_No_Bareek', 'Local_Kraft', 'Local_Kraft_Bareek', 'Imported_Kraft', 'Imported_Kraft_Bareek', 'Super_Fluting', 'Super_Fluting_Bareek']
 
                sheets=[]
                for i in clientssheets.keys():
                    if i not in toremove:
                        sheets.append(i)
                datas=pd.read_excel(r'book.xlsx' , index_col=None,sheet_name=sheets, usecols=['DATE','RECIEPT_NUMBER','CLIENT_ID','CLIENT_NAME','CONTACT_NO','DETAILS_OF_BILL','DEBIT','CREDIT','CREDIT_DETAILS','RENT','BALANCE'])

                rcp_df=pd.concat(datas[frame] for frame in datas.keys())
                w=rcp_df['RECIEPT_NUMBER'].unique()
                if(len(w)!=0):
                    w[::-1].sort()
                    self.rcp_no.setText(str(w[0]+1)) 

                else:
                    self.rcp_no.setText(str(1))
                    
                lis=[]
                slis=[]

                
                number_of_rows = self.cashbillDetailstable.rowCount()
                number_of_columns = self.cashbillDetailstable.columnCount()
                for i in range(number_of_rows):
                    lis=[]
                    for j in range(number_of_columns):
                        x=str(self.cashbillDetailstable.item(i, j).text()) 
                        x.replace('Tota_Packets','Packets')
                        x.replace('Reel_Packets','Packets')
                        list1=[]
                        list1=x.split()
                        if(list1[-1]=='Packets'):
                            del list1[2] 
                            del list1[2]
                        strings=(' ').join(list1)
                        lis.append(strings)
                          
                    slis.append(lis)
                         
                
                df = pd.DataFrame(slis, columns =['Details', 'Price'])
                         
                
                RECIEPT_NUMBER = int(self.rcp_no.toPlainText().strip())
                DEBIT = float(self.totalTextField.toPlainText().strip()) + float(self.RentTextField.toPlainText().strip())
                RENT = self.RentTextField.toPlainText().strip()
                CREDIT = self.CreditTextField.toPlainText().strip()
                
                CREDIT_DETAILS = self.CreditDetailsTextField.toPlainText().strip()
                df = pd.DataFrame(slis, columns =['Details', 'Price'])
                
                pdf_date = DATE
                pdf_client_name = CLIENT_NAME
                pdf_client_contact_number = CONTACT_NO
                pdf_reciept_number = RECIEPT_NUMBER
                pdf_debit = DEBIT
                pdf_credit = CREDIT
                pdf_credit_details = CREDIT_DETAILS
                pdf_rent = RENT
                pdf_previous_balance = self.prevbalance.toPlainText()
                pdf_previous_bill_number = self.prevrcp.toPlainText()
                pdf_total_balance = BALANCE
                pdf_dataframe = df.copy()
                pdf_dataframe['Details']=pdf_dataframe['Details'].str.replace('Tota_Packets','Packets')
                pdf_dataframe['Details']=pdf_dataframe['Details'].str.replace('Reel_Packets','Packets') 
                write_excel(CashCustomers,text,r'book.xlsx' )
                write_excel( self.Fluting, "Fluting",r'book.xlsx' )
                write_excel( self.Fluting_bareek, "Fluting_Bareek",r'book.xlsx')   
                write_excel( self.L1, "L1",r'book.xlsx' )
                write_excel(self.L1_bareek , "L1_Bareek",r'book.xlsx')
                write_excel( self.L2, "L2",r'book.xlsx' )
                write_excel( self.L2_Bareek, "L2_Bareek",r'book.xlsx')
                write_excel(self.testliner , "Test_Liner",r'book.xlsx' )
                write_excel( self.testliner_bareek, "Test_Liner_Bareek",r'book.xlsx')
                write_excel( self.boxboard2_5, "Box_Board_2_5_No",r'book.xlsx' )
                write_excel( self.boxboard2_5_bareek, "Box_Board_2_5_No_Bareek",r'book.xlsx')
                write_excel( self.boxboard3, "Box_Board_3_No",r'book.xlsx' )
                write_excel( self.boxboard3_bareek, "Box_Board_3_No_Bareek",r'book.xlsx')
                write_excel(self.localkraft , "Local_Kraft",r'book.xlsx' )
                write_excel( self.localkraft_bareek, "Local_Kraft_Bareek",r'book.xlsx')
                write_excel(self.importedkraft , "Imported_Kraft",r'book.xlsx' )
                write_excel( self.importedkraft_bareek, "Imported_Kraft_Bareek",r'book.xlsx')
                write_excel( self.Super_Fluting, "Super_Fluting",r'book.xlsx' )
                write_excel( self.Super_Fluting_bareek, "Super_Fluting_Bareek",r'book.xlsx')
                write_excel( self.reelsstock, "reels_stock",r'book.xlsx')
                write_excel( self.totaystock, "totay",r'book.xlsx')
                write_excel( self.stock_out_totay, "tota_stock_in_out",r'book.xlsx')
                write_excel( self.stock_out_reels, "reels_stock_in_out",r'book.xlsx')
                write_excel( self.stock_out_rolls, "rolls_stock_in_out",r'book.xlsx')
                msg = QMessageBox()  # create an instance of it
                msg.setIcon(QMessageBox.Information)  # set icon
                msg.setText("Entry Saved Successfully")  # set text
                
                msg.setWindowTitle("Message")  # set title
                message = msg.exec_()
                 
                
                
                
                '''super().Ui_Form.()      '''
                #super(Ui_MainWindow_client, self).updatetable()
                
                
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Question) 
                msgBox.setWindowIcon(QtGui.QIcon("whatsapp-logo.png"))
                msgBox.setText("Do you want to send this bill to Client's WhatsApp?")  # set text   
                msgBox.setWindowTitle("WhatsApp Message Send Option")  
                msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                
                returnValue = msgBox.exec()
                if returnValue == QMessageBox.Ok:
                        whtsapp(self.contact.toPlainText().strip(),CLIENT_NAME)
                
                
                
                self.cashbillDetailstable.setRowCount(0)
                self.itemListComboBox.setCurrentIndex(0)
                self.totalTextField.setText("0")
                self.m_ledger.setText( str(ledger) )
                self.prevbalance.setText( str(BALANCE) )
                self.prevrcp.setText( str(RECIEPT_NUMBER) )  
                self.RentTextField.setText('0')
                self.CreditTextField.setText('0')
                self.CreditDetailsTextField.setText('')
                
                self.itemListComboBox.setCurrentIndex(0)
                self.reset_reels_container()
                self.reset_packets_container()
                self.reset_jutta_container()
                self.reset_totay_container()
                self.reset_raddi_container()
                self.reset_nali_container()
                self.reset_rolls_container()
                self.RollsContainer.setVisible(False)
                self.ReelsContainer.setVisible(False)
                self.PacketsContainer.setVisible(False)
                self.JuttaContainer.setVisible(False)
                self.RaddiContainer.setVisible(False)
                self.TotayContainer.setVisible(False)
                self.NaliContainer.setVisible(False)
                MainWindow.showMaximized() 
        
        
        def generate_cashbill_pdf():
            global pdf_date
            global pdf_client_name
            global pdf_client_contact_number
            global pdf_reciept_number
            global pdf_debit
            global pdf_credit
            global pdf_credit_details
            global pdf_rent
            global pdf_previous_balance
            global pdf_previous_bill_number
            global pdf_total_balance
            global pdf_dataframe
            
            # Rent
            # Credit Details
            # Credit
            # Debit
            # Client Name
            # Reciept Number
            # Contact Number
            # Date
            # Previous Balance
            # Previous Bill Number
            # Total Balance
            
            other_list=[]
            other_list.append(pdf_rent)
            other_list.append(pdf_credit_details)
            other_list.append(pdf_credit)
            other_list.append(pdf_debit)
            other_list.append(pdf_client_name)
            other_list.append(pdf_reciept_number)
            other_list.append(pdf_client_contact_number)
            other_list.append(pdf_date)
            other_list.append(pdf_previous_balance)
            other_list.append(pdf_previous_bill_number)
            other_list.append(pdf_total_balance)
            
            
            #rowscount=self.cashbillDetailstable.rowCount()
            #headercount =  self.cashbillDetailstable.columnCount()
            #table=pd.DataFrame(columns=[self.cashbillDetailstable.horizontalHeaderItem(i).text() for i in range(headercount) ],
            #                index=[x for x in range(rowscount)])
            #for row in range(rowscount):
            #    for col in range(headercount):
            #        headertext =  self.cashbillDetailstable.horizontalHeaderItem(col).text()
            #        cell =  self.cashbillDetailstable.item(row, col).text()  # get cell at row, col
            #        table[headertext][row]=cell
            #table.to_csv('C:\\Users\\abeer\Downloads\\Factory_management_system\\table.csv')
            #table.to_csv('table.csv')
            #print('\n\nList',mylist)
            if(len(pdf_dataframe)==0):
                gc.collect()   
                close_window()
            else:
                if (len(error) != 0):
                    #print(errors, " erors ")
                    msg = QMessageBox()  # create an instance of it
                    msg.setIcon(QMessageBox.Information)  # set icon
                    msgs = "Cannot Print Bill As there are errors"
                    msg.setText(msgs)  # set text

                    '''msg.setInformativeText()'''  # set information under the main text
                    msg.setWindowTitle("Alert")  # set title
                    message = msg.exec_()
                    error = []
                else:
                    generate_customer_invoice(pdf_dataframe,other_list)
                    msgBox = QMessageBox()
                    msgBox.setIcon(QMessageBox.Question) 
                    msgBox.setWindowIcon(QtGui.QIcon("whatsapp-logo.png"))
                    msgBox.setText("Do you want to send this bill to Client's WhatsApp?")  # set text   
                    msgBox.setWindowTitle("WhatsApp Message Send Option")  
                    msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                        
                    returnValue = msgBox.exec()
                    if returnValue == QMessageBox.Ok:
                            whtsapp(pdf_client_contact_number,pdf_reciept_number,pdf_client_name)
                    pass
                gc.collect()   
                close_window()
            

        self.CheckOutButton.clicked.connect(cashbill) #8 spaces position ... outside cashbill func
        self.CheckOutButton.clicked.connect(generate_cashbill_pdf) 
    
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi6(Form)
    Form.showMaximized()
    sys.exit(app.exec_())
    

